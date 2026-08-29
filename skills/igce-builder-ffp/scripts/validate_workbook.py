#!/usr/bin/env python3
"""Validate an FFP workbook structurally and with optional engine execution."""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from recompute_expected_values import InputError, calculate, load_payload


DEFAULT_SHEETS = [
    "IGCE Summary",
    "Cost Buildup",
    "Scenario Analysis",
    "Rate Validation",
    "Travel Detail",
    "Methodology",
    "Raw Data",
]

CELL_REF = re.compile(
    r"^(?:'((?:[^']|'')+)'|([^!]+))!\$?([A-Za-z]{1,3})\$?([1-9][0-9]*)$"
)
COST_B_REF = re.compile(
    r"(?:'Cost Buildup'|Cost Buildup)!\$?B\$?([1-9][0-9]*)",
    re.IGNORECASE,
)


def normalize_formula(value: Any) -> str:
    return re.sub(r"\s+", "", str(value)).upper()


def parse_cell_ref(reference: str) -> tuple[str, str]:
    match = CELL_REF.fullmatch(reference.strip())
    if not match:
        raise InputError(f"invalid workbook cell reference: {reference}")
    sheet = (match.group(1) or match.group(2)).replace("''", "'")
    coordinate = f"{match.group(3).upper()}{match.group(4)}"
    return sheet, coordinate


def cell_value(workbook: Any, reference: str) -> Any:
    sheet, coordinate = parse_cell_ref(reference)
    if sheet not in workbook.sheetnames:
        raise InputError(f"cell reference uses missing sheet: {reference}")
    return workbook[sheet][coordinate].value


def check_formula(
    failures: list[str],
    sheet: Any,
    coordinate: str,
    *,
    expected: str | None = None,
    contains: list[str] | None = None,
    not_contains: list[str] | None = None,
) -> None:
    value = sheet[coordinate].value
    if not isinstance(value, str) or not value.startswith("="):
        failures.append(f"{sheet.title}!{coordinate} is not a formula")
        return
    normalized = normalize_formula(value)
    if expected is not None and normalized != normalize_formula(expected):
        failures.append(f"{sheet.title}!{coordinate} formula does not match expected structure")
    for item in contains or []:
        if normalize_formula(item) not in normalized:
            failures.append(f"{sheet.title}!{coordinate} formula is missing {item}")
    for item in not_contains or []:
        if normalize_formula(item) in normalized:
            failures.append(f"{sheet.title}!{coordinate} formula contains forbidden text {item}")


# --- Rendered-text clipping audit -------------------------------------------
# A text cell overflows into the next cell only when that neighbour is empty.
# When the neighbour is occupied the label is cut off in the printed workbook,
# so every such label must fit its column, wrap, or be merged across the block.

CLIPPING_GLYPH_WIDTHS = {
    " ": 0.45,
    ".": 0.45,
    ",": 0.45,
    ";": 0.45,
    ":": 0.45,
    "'": 0.35,
    "`": 0.45,
    "!": 0.45,
    "|": 0.45,
    "(": 0.55,
    ")": 0.55,
    "[": 0.55,
    "]": 0.55,
    "{": 0.60,
    "}": 0.60,
    "-": 0.60,
    "/": 0.55,
    "\\": 0.55,
    '"': 0.60,
    "%": 1.50,
    "@": 1.70,
    "$": 1.00,
}
CLIPPING_LOWER_NARROW = "ijl"
CLIPPING_LOWER_SEMI = "frt"
CLIPPING_LOWER_WIDE = "mw"
CLIPPING_UPPER_NARROW = "I"
CLIPPING_UPPER_WIDE = "MW"
CLIPPING_BOLD_FACTOR = 1.14
CLIPPING_ABSOLUTE_TOLERANCE = 0.75
CLIPPING_RELATIVE_TOLERANCE = 0.04
CLIPPING_DEFAULT_WIDTH = 8.43
CLIPPING_MESSAGE_TEXT_LIMIT = 120


def glyph_width(character: str) -> float:
    """Width of one glyph in Excel column-width units (1.0 = one digit)."""
    if character in CLIPPING_GLYPH_WIDTHS:
        return CLIPPING_GLYPH_WIDTHS[character]
    if character in CLIPPING_LOWER_NARROW:
        return 0.48
    if character in CLIPPING_LOWER_SEMI:
        return 0.59
    if character in CLIPPING_LOWER_WIDE:
        return 1.66
    if character in CLIPPING_UPPER_NARROW:
        return 0.45
    if character in CLIPPING_UPPER_WIDE:
        return 1.55
    if character.islower():
        return 0.96
    if character.isupper():
        return 1.05
    return 1.0


def estimated_text_width(text: str, font: Any = None) -> float:
    """Estimated rendered width of a label in column-width units."""
    lines = str(text).split("\n")
    units = max((sum(glyph_width(character) for character in line) for line in lines), default=0.0)
    size = getattr(font, "size", None) or 11.0
    if float(size) != 11.0:
        units *= float(size) / 11.0
    if getattr(font, "bold", False):
        units *= CLIPPING_BOLD_FACTOR
    return units


def column_width_map(sheet: Any) -> tuple[dict[int, float], float]:
    """Explicit column widths by index plus the sheet default width."""
    widths: dict[int, float] = {}
    for letter, dimension in sheet.column_dimensions.items():
        if dimension.width is None:
            continue
        # In-memory dimensions created by a generator carry no min/max, so fall
        # back to the column the dimension is keyed under.
        try:
            own = column_index_from_string(letter)
        except ValueError:
            own = None
        first = dimension.min or own or 1
        last = dimension.max or own or first
        for index in range(first, last + 1):
            widths[index] = float(dimension.width)
    default = sheet.sheet_format.defaultColWidth or CLIPPING_DEFAULT_WIDTH
    return widths, float(default)


def merged_ranges_by_anchor(sheet: Any) -> tuple[dict[tuple[int, int], Any], set[tuple[int, int]]]:
    """Merged-range lookup keyed by anchor cell, plus every covered cell."""
    anchors: dict[tuple[int, int], Any] = {}
    covered: set[tuple[int, int]] = set()
    for merged in sheet.merged_cells.ranges:
        anchors[(merged.min_row, merged.min_col)] = merged
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                covered.add((row, column))
    return anchors, covered


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    return isinstance(value, str) and not value.strip()


def text_clipping_audit(
    workbook: Any,
    *,
    sheets: list[str] | None = None,
    exempt: set[str] | None = None,
) -> list[str]:
    """Flag text that is cut off in print because an occupied neighbour blocks overflow."""
    failures: list[str] = []
    skipped = exempt or set()
    for sheet in workbook.worksheets:
        if sheets is not None and sheet.title not in sheets:
            continue
        widths, default_width = column_width_map(sheet)
        anchors, covered = merged_ranges_by_anchor(sheet)
        max_column = sheet.max_column
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.strip() or value.startswith("="):
                    continue
                position = (cell.row, cell.column)
                if position in covered and position not in anchors:
                    continue
                alignment = cell.alignment
                if alignment.wrap_text or alignment.horizontal in {"fill", "distributed"}:
                    continue
                if f"{sheet.title}!{cell.coordinate}" in skipped:
                    continue
                merged = anchors.get(position)
                first_column = cell.column
                last_column = merged.max_col if merged is not None else cell.column
                available = sum(
                    widths.get(index, default_width)
                    for index in range(first_column, last_column + 1)
                )
                needed = estimated_text_width(value, cell.font)
                tolerance = max(
                    CLIPPING_ABSOLUTE_TOLERANCE,
                    CLIPPING_RELATIVE_TOLERANCE * available,
                )
                if needed <= available + tolerance:
                    continue
                blockers = []
                if alignment.horizontal in {"right", "center", "centerContinuous"}:
                    left = first_column - 1
                    if left < 1:
                        blockers.append("the left sheet edge")
                    elif not _is_blank(sheet.cell(row=cell.row, column=left).value):
                        blockers.append(sheet.cell(row=cell.row, column=left).coordinate)
                if alignment.horizontal != "right":
                    right = last_column + 1
                    if right <= max_column and not _is_blank(
                        sheet.cell(row=cell.row, column=right).value
                    ):
                        blockers.append(sheet.cell(row=cell.row, column=right).coordinate)
                if not blockers:
                    continue
                shown = value.strip()
                if len(shown) > CLIPPING_MESSAGE_TEXT_LIMIT:
                    shown = shown[: CLIPPING_MESSAGE_TEXT_LIMIT - 3] + "..."
                span = (
                    cell.column_letter
                    if merged is None
                    else f"{cell.column_letter}:{get_column_letter(last_column)}"
                )
                failures.append(
                    f"{sheet.title}!{cell.coordinate} is clipped in print: '{shown}' needs about "
                    f"{needed:.1f} column-width units but column {span} gives {available:g} and "
                    f"{', '.join(blockers)} blocks the overflow; widen the column to at least "
                    f"{math.ceil(needed):g}, enable wrap text with adequate row height, merge the "
                    f"label across the block, or shorten it"
                )
    return failures


def structural_audit(workbook: Any, payload: dict[str, Any]) -> list[str]:
    failures: list[str] = []
    required_sheets = payload.get("required_sheets", DEFAULT_SHEETS)
    if not isinstance(required_sheets, list) or not all(
        isinstance(item, str) for item in required_sheets
    ):
        raise InputError("required_sheets must be an array of strings")
    for sheet_name in required_sheets:
        if sheet_name not in workbook.sheetnames:
            failures.append(f"missing required sheet: {sheet_name}")

    if "IGCE Summary" in workbook.sheetnames:
        summary = workbook["IGCE Summary"]
        check_formula(
            failures,
            summary,
            "B11",
            contains=["VALUE(LEFT(B10,4))", "VALUE(MID(B10,6,2))"],
            not_contains=["DATEDIF", "YEAR("],
        )
        check_formula(
            failures,
            summary,
            "B12",
            contains=["B6", "B11", "^"],
        )
        if summary["B12"].number_format != "0.0000":
            failures.append("IGCE Summary!B12 must display the aging factor as 0.0000")

    formula_count = 0
    formula_error_tokens = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    upper = value.upper()
                    if any(token in upper for token in formula_error_tokens):
                        failures.append(f"{sheet.title}!{cell.coordinate} contains a formula error token")
                elif isinstance(value, str) and value[:1] in {"+", "-", "@"}:
                    failures.append(
                        f"{sheet.title}!{cell.coordinate} starts with a formula-trigger character"
                    )
    if formula_count == 0:
        failures.append("workbook contains no formulas")

    if "Cost Buildup" in workbook.sheetnames:
        buildup = workbook["Cost Buildup"]
        starts: list[int] = []
        for row_index in range(1, buildup.max_row + 1):
            label = buildup.cell(row_index, 1).value
            if isinstance(label, str) and label.startswith("Cost Buildup:"):
                starts.append(row_index)
        if not starts:
            failures.append("Cost Buildup contains no recognized labor blocks")
        for block_index, base in enumerate(starts):
            expected_base = 1 + block_index * 19
            if base != expected_base:
                failures.append(
                    f"Cost Buildup block {block_index + 1} starts at row {base}, expected {expected_base}"
                )
            if buildup[f"B{base + 5}"].value is not None:
                failures.append(f"Cost Buildup!B{base + 5} must be the blank separator row")
            check_formula(
                failures,
                buildup,
                f"B{base + 2}",
                contains=["'IGCE SUMMARY'!$B$12"],
            )
            if buildup[f"B{base + 2}"].number_format != "0.0000":
                failures.append(
                    f"Cost Buildup!B{base + 2} must display the aging factor as 0.0000"
                )
            formulas = {
                base + 3: f"=B{base + 1}*B{base + 2}",
                base + 4: f"=B{base + 3}/2080",
                base + 6: "='IGCE Summary'!$B$2",
                base + 7: f"=B{base + 4}*B{base + 6}",
                base + 8: f"=B{base + 4}+B{base + 7}",
                base + 9: "='IGCE Summary'!$B$3",
                base + 10: f"=B{base + 8}*B{base + 9}",
                base + 11: f"=B{base + 8}+B{base + 10}",
                base + 12: "='IGCE Summary'!$B$4",
                base + 13: f"=B{base + 11}*B{base + 12}",
                base + 14: f"=B{base + 11}+B{base + 13}",
                base + 15: "='IGCE Summary'!$B$5",
                base + 16: f"=B{base + 14}*B{base + 15}",
                base + 17: f"=B{base + 14}+B{base + 16}",
                base + 18: f"=B{base + 17}/B{base + 4}",
            }
            for row_number, expected_formula in formulas.items():
                check_formula(
                    failures,
                    buildup,
                    f"B{row_number}",
                    expected=expected_formula,
                )

    for sheet_name in ("IGCE Summary", "Scenario Analysis", "Rate Validation"):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                for match in COST_B_REF.finditer(value):
                    referenced_row = int(match.group(1))
                    if (referenced_row - 4) % 19 == 0:
                        failures.append(
                            f"{sheet_name}!{cell.coordinate} uses Cost Buildup row {referenced_row} "
                            "as a cross-sheet input; that row is Aged Annual Wage"
                        )

    assertions = payload.get("formula_assertions", [])
    if not isinstance(assertions, list):
        raise InputError("formula_assertions must be an array")
    for index, assertion in enumerate(assertions):
        if not isinstance(assertion, dict) or not isinstance(assertion.get("cell"), str):
            raise InputError(f"formula_assertions[{index}] must contain a cell string")
        sheet_name, coordinate = parse_cell_ref(assertion["cell"])
        if sheet_name not in workbook.sheetnames:
            failures.append(f"formula assertion references missing sheet: {sheet_name}")
            continue
        contains = assertion.get("contains", [])
        not_contains = assertion.get("not_contains", [])
        if not isinstance(contains, list) or not all(isinstance(item, str) for item in contains):
            raise InputError(f"formula_assertions[{index}].contains must be an array of strings")
        if not isinstance(not_contains, list) or not all(
            isinstance(item, str) for item in not_contains
        ):
            raise InputError(
                f"formula_assertions[{index}].not_contains must be an array of strings"
            )
        equals = assertion.get("equals")
        if equals is not None and not isinstance(equals, str):
            raise InputError(f"formula_assertions[{index}].equals must be a string")
        check_formula(
            failures,
            workbook[sheet_name],
            coordinate,
            expected=equals,
            contains=contains,
            not_contains=not_contains,
        )
    failures.extend(text_clipping_audit(workbook))
    return failures


def find_soffice() -> Path | None:
    discovered = shutil.which("soffice")
    if discovered:
        return Path(discovered)
    mac_path = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    return mac_path if mac_path.is_file() else None


def recalculate_with_libreoffice(source: Path, executable: Path) -> tuple[tempfile.TemporaryDirectory[str], Path]:
    temp = tempfile.TemporaryDirectory(prefix="ffp-workbook-validation-")
    temp_root = Path(temp.name)
    input_dir = temp_root / "input"
    output_dir = temp_root / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    copied = input_dir / source.name
    shutil.copy2(source, copied)
    command = [
        str(executable),
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(output_dir),
        str(copied),
    ]
    completed = subprocess.run(command, capture_output=True, text=True, timeout=120, check=False)
    recalculated = output_dir / source.name
    if completed.returncode != 0 or not recalculated.is_file():
        temp.cleanup()
        detail = (completed.stderr or completed.stdout).strip()
        raise InputError(f"LibreOffice recalculation failed: {detail or 'no output file'}")
    return temp, recalculated


def close_enough(actual: float, expected: float, tolerance: float) -> bool:
    return math.isclose(actual, expected, rel_tol=tolerance, abs_tol=max(0.01, tolerance))


def compare_results(
    workbook: Any,
    expected: dict[str, Any],
    tolerance: float,
) -> list[str]:
    failures: list[str] = []

    def compare(reference: str, target: float, label: str) -> None:
        try:
            value = cell_value(workbook, reference)
        except InputError as exc:
            failures.append(str(exc))
            return
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            failures.append(f"{label} at {reference} has no calculated numeric value")
            return
        if not close_enough(float(value), target, tolerance):
            failures.append(
                f"{label} at {reference} is {float(value):.6f}, expected {target:.6f}"
            )

    for line in expected["labor_lines"]:
        if "workbook_fbr_cell" in line:
            compare(
                line["workbook_fbr_cell"],
                line["fully_burdened_rate"],
                f"{line['name']} FBR",
            )
        if "workbook_total_cell" in line:
            compare(
                line["workbook_total_cell"],
                line["labor_total"],
                f"{line['name']} labor total",
            )
    for line in expected["non_labor_lines"]:
        if "workbook_total_cell" in line:
            compare(
                line["workbook_total_cell"],
                line["amount"],
                f"{line['name']} amount",
            )
    if "workbook_grand_total_cell" in expected:
        compare(
            expected["workbook_grand_total_cell"],
            expected["grand_total"],
            "grand total",
        )
    return failures


def cached_error_audit(workbook: Any) -> list[str]:
    failures: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("#"):
                    failures.append(f"{sheet.title}!{cell.coordinate} has cached error {value}")
    return failures


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Audit an FFP workbook and optionally verify formula execution with LibreOffice."
    )
    parser.add_argument("workbook", type=Path, help="Workbook to validate")
    parser.add_argument("--expected", required=True, type=Path, help="Validation-input JSON")
    parser.add_argument(
        "--engine",
        choices=("auto", "none", "libreoffice"),
        default="auto",
        help="Formula execution engine policy",
    )
    parser.add_argument("--tolerance", type=float, default=0.01, help="Relative comparison tolerance")
    parser.add_argument("--json", action="store_true", help="Emit a JSON result")
    args = parser.parse_args()

    if not math.isfinite(args.tolerance) or args.tolerance < 0:
        print("ERROR: tolerance must be finite and non-negative", file=sys.stderr)
        return 2
    if not args.workbook.is_file():
        print(f"ERROR: workbook not found: {args.workbook}", file=sys.stderr)
        return 2

    try:
        payload = load_payload(args.expected)
        expected = calculate(payload)
        formula_workbook = load_workbook(args.workbook, data_only=False)
        structural_failures = structural_audit(formula_workbook, payload)
        failures = list(structural_failures)
    except (InputError, OSError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    engine_used: str | None = None
    engine_note: str
    temp: tempfile.TemporaryDirectory[str] | None = None
    try:
        if args.engine == "none":
            engine_note = "Formula execution was not requested."
        else:
            soffice = find_soffice()
            if soffice is None:
                if args.engine == "libreoffice":
                    failures.append("LibreOffice was required but soffice was not found")
                    engine_note = "LibreOffice was required but unavailable."
                else:
                    engine_note = "Formula execution was not independently verified in Excel or LibreOffice."
            else:
                temp, recalculated = recalculate_with_libreoffice(args.workbook, soffice)
                calculated_workbook = load_workbook(recalculated, data_only=True)
                failures.extend(cached_error_audit(calculated_workbook))
                failures.extend(compare_results(calculated_workbook, expected, args.tolerance))
                engine_used = "LibreOffice"
                engine_note = "LibreOffice formula execution and cached-value comparison ran."
    except (InputError, OSError, subprocess.SubprocessError, ValueError) as exc:
        failures.append(str(exc))
        engine_note = "LibreOffice execution failed."
    finally:
        if temp is not None:
            temp.cleanup()

    result = {
        "status": "pass" if not failures else "fail",
        "formula_structure": "pass" if not structural_failures else "fail",
        "independent_recomputation": "pass",
        "independent_grand_total": expected["grand_total"],
        "engine": engine_used,
        "engine_note": engine_note,
        "failures": failures,
    }
    if args.json:
        print(json.dumps(result, indent=2, sort_keys=True))
    else:
        if failures:
            print("VALIDATION FAILED")
            for failure in failures:
                print(f"- {failure}")
        elif engine_used:
            print("Formula structure, independent calculations, and LibreOffice formula execution passed.")
        else:
            print(
                "Formula structure and independent calculations passed. "
                "Formula execution was not independently verified in Excel or LibreOffice."
            )
        print(f"Independent grand total: {expected['grand_total']:.2f}")
        print(engine_note)
    return 0 if not failures else 1


if __name__ == "__main__":
    raise SystemExit(main())
