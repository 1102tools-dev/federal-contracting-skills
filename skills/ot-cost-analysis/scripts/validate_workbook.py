#!/usr/bin/env python3
"""Validate OT workbook structure and compare calculated values."""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import subprocess
import sys
import tempfile
import zipfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string, get_column_letter

from recompute_expected_values import InputError, calculate, load_payload


REQUIRED_SHEETS = [
    "OT Cost Summary",
    "Milestone Detail",
    "Scenario Analysis",
    "Labor Benchmarking",
    "Cost Share & Funding",
    "Methodology",
    "Raw Data",
]
CELL_REF = re.compile(r"^(?:'((?:[^']|'')+)'|([^!]+))!\$?([A-Za-z]{1,3})\$?([1-9][0-9]*)$")


def normalize_formula(value: Any) -> str:
    return re.sub(r"\s+", "", str(value)).upper()


def parse_cell_ref(reference: str) -> tuple[str, str]:
    match = CELL_REF.fullmatch(reference.strip())
    if not match:
        raise InputError(f"invalid workbook cell reference: {reference}")
    sheet = (match.group(1) or match.group(2)).replace("''", "'")
    return sheet, f"{match.group(3).upper()}{match.group(4)}"


def value_at(workbook: Any, reference: str) -> Any:
    sheet, coordinate = parse_cell_ref(reference)
    if sheet not in workbook.sheetnames:
        raise InputError(f"cell reference uses missing sheet: {reference}")
    return workbook[sheet][coordinate].value


def check_formula(
    failures: list[str],
    workbook: Any,
    reference: str,
    *,
    expected: str | None = None,
    contains: list[str] | None = None,
    not_contains: list[str] | None = None,
) -> None:
    try:
        value = value_at(workbook, reference)
    except InputError as exc:
        failures.append(str(exc))
        return
    if not isinstance(value, str) or not value.startswith("="):
        failures.append(f"{reference} is not a formula")
        return
    normalized = normalize_formula(value)
    if expected is not None and normalized != normalize_formula(expected):
        failures.append(f"{reference} formula does not match expected structure")
    for item in contains or []:
        if normalize_formula(item) not in normalized:
            failures.append(f"{reference} formula is missing {item}")
    for item in not_contains or []:
        if normalize_formula(item) in normalized:
            failures.append(f"{reference} formula contains forbidden text {item}")


BENCHMARK_REF = re.compile(r"'?LABOR\s?BENCHMARKING'?!\$?[A-Z]{1,3}\$?([1-9][0-9]*)", re.I)
NARRATIVE_HEADERS = {"description", "basis", "source note"}
MIN_NARRATIVE_WIDTH = 28.0


def _row_strings(row: Any) -> list[str]:
    return [
        cell.value
        for cell in row
        if isinstance(cell.value, str) and not cell.value.startswith("=")
    ]


def _row_benchmark_row_number(row: Any) -> int | None:
    for cell in row:
        value = cell.value
        if isinstance(value, str) and value.startswith("="):
            match = BENCHMARK_REF.search(value)
            if match:
                return int(match.group(1))
    return None


def labor_benchmark_audit(workbook: Any) -> list[str]:
    """Every priced labor line matches its own benchmark row or names a proxy."""
    failures: list[str] = []
    if "Milestone Detail" not in workbook.sheetnames or "Labor Benchmarking" not in workbook.sheetnames:
        return failures
    benchmarks = workbook["Labor Benchmarking"]
    for row in workbook["Milestone Detail"].iter_rows():
        benchmark_row = _row_benchmark_row_number(row)
        if benchmark_row is None:
            continue
        texts = _row_strings(row)
        if not texts:
            continue
        category = texts[0].strip()
        benchmark_value = benchmarks.cell(row=benchmark_row, column=1).value
        benchmark_name = benchmark_value.strip() if isinstance(benchmark_value, str) else ""
        lowered_category = category.lower()
        lowered_benchmark = benchmark_name.lower()
        matched = bool(lowered_benchmark) and (
            lowered_category in lowered_benchmark or lowered_benchmark in lowered_category
        )
        has_proxy = any("proxy" in text.lower() for text in texts[1:])
        if not matched and not has_proxy:
            failures.append(
                f"Milestone Detail!{row[0].coordinate} prices '{category}' from the benchmark row for "
                f"'{benchmark_name or 'unknown'}' without its own benchmark row or a Basis naming the proxy source"
            )
    return failures


def hours_reconciliation_audit(workbook: Any) -> list[str]:
    """Identical per-category hours across milestones need an hours-basis note."""
    failures: list[str] = []
    if "Milestone Detail" not in workbook.sheetnames:
        return failures
    detail = workbook["Milestone Detail"]
    has_note = any(
        isinstance(cell.value, str) and "hours basis" in cell.value.lower()
        for row in detail.iter_rows()
        for cell in row
    )
    if has_note:
        return failures
    hours_by_category: dict[str, list[float]] = {}
    for row in detail.iter_rows():
        if _row_benchmark_row_number(row) is None:
            continue
        texts = _row_strings(row)
        numbers = [
            float(cell.value)
            for cell in row
            if isinstance(cell.value, (int, float)) and not isinstance(cell.value, bool)
        ]
        if not texts or len(numbers) != 1:
            continue
        hours_by_category.setdefault(texts[0].strip().lower(), []).append(numbers[0])
    for category, hours in sorted(hours_by_category.items()):
        if len(hours) >= 2 and len(set(hours)) == 1:
            failures.append(
                f"Milestone Detail repeats identical hours ({hours[0]:g}) for '{category}' across "
                f"{len(hours)} milestones with no hours-basis note reconciling hours to duration and staffing"
            )
    return failures


def narrative_format_audit(workbook: Any, reported: set[str] | None = None) -> list[str]:
    """Narrative columns must wrap text and meet the width floor.

    Cells reported here are recorded in ``reported`` so the clipping audit does
    not raise a second finding against the same unwrapped narrative cell.
    """
    failures: list[str] = []
    seen: set[tuple[str, str]] = set()
    for sheet_name in ("OT Cost Summary", "Milestone Detail"):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if not isinstance(cell.value, str):
                    continue
                header = re.sub(r"\s+", " ", cell.value).strip().lower()
                if header not in NARRATIVE_HEADERS:
                    continue
                letter = cell.column_letter
                if (sheet_name, letter) in seen:
                    continue
                seen.add((sheet_name, letter))
                width = sheet.column_dimensions[letter].width
                if width is None or width < MIN_NARRATIVE_WIDTH:
                    failures.append(
                        f"{sheet_name} narrative column {letter} ('{cell.value.strip()}') width "
                        f"{width or 0:g} is below the {MIN_NARRATIVE_WIDTH:g} floor"
                    )
                for row_number in range(cell.row + 1, sheet.max_row + 1):
                    below = sheet.cell(row=row_number, column=cell.column)
                    value = below.value
                    if (
                        isinstance(value, str)
                        and value.strip()
                        and not value.startswith("=")
                        and not below.alignment.wrap_text
                    ):
                        failures.append(
                            f"{sheet_name}!{below.coordinate} narrative cell under "
                            f"'{cell.value.strip()}' does not have wrap text enabled"
                        )
                        if reported is not None:
                            reported.add(f"{sheet_name}!{below.coordinate}")
                        break
    return failures


MAX_PORTRAIT_COLUMNS = 8


def _used_column_count(sheet: Any) -> int:
    widest = 0
    for row in sheet.iter_rows():
        for cell in row:
            if cell.value is not None and cell.column > widest:
                widest = cell.column
    return widest


def print_setup_audit(workbook: Any) -> list[str]:
    """Every canonical sheet must print cleanly: print area, fitToPage, landscape when wide."""
    failures: list[str] = []
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        if not sheet.print_area:
            failures.append(f"{sheet_name} has no print area set over the populated range")
        page_setup_properties = sheet.sheet_properties.pageSetUpPr
        if page_setup_properties is None or not page_setup_properties.fitToPage:
            failures.append(f"{sheet_name} does not enable fitToPage scaling")
        used_columns = _used_column_count(sheet)
        if used_columns > MAX_PORTRAIT_COLUMNS and sheet.page_setup.orientation != "landscape":
            failures.append(
                f"{sheet_name} uses {used_columns} columns in portrait orientation; sheets wider "
                f"than {MAX_PORTRAIT_COLUMNS} used columns must print landscape"
            )
    return failures


def is_recost_workbook(workbook: Any) -> bool:
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows(min_row=1, max_row=3):
            for cell in row:
                if isinstance(cell.value, str) and "recost" in cell.value.lower():
                    return True
    return False


def _header_map(row: Any) -> dict[str, int]:
    return {
        re.sub(r"\s+", " ", cell.value).strip().lower(): cell.column
        for cell in row
        if isinstance(cell.value, str)
    }


def recost_audit(workbook: Any, payload: dict[str, Any]) -> list[str]:
    """Recost-specific gates: no orphan benchmarks, decomposed labor deltas, register coverage."""
    failures: list[str] = []
    if "Milestone Detail" not in workbook.sheetnames:
        return failures
    detail = workbook["Milestone Detail"]
    detail_text = "\n".join(
        cell.value.lower()
        for row in detail.iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    )

    if "Labor Benchmarking" in workbook.sheetnames:
        benchmarks = workbook["Labor Benchmarking"]
        header_columns: dict[str, int] = {}
        header_row_number = 0
        for row in benchmarks.iter_rows():
            columns = _header_map(row)
            if "role" in columns or "labor category" in columns:
                header_columns = columns
                header_row_number = row[0].row
                break
        role_column = header_columns.get("role") or header_columns.get("labor category")
        if role_column:
            for row_number in range(header_row_number + 1, benchmarks.max_row + 1):
                value = benchmarks.cell(row=row_number, column=role_column).value
                if not isinstance(value, str) or not value.strip():
                    continue
                role = value.strip()
                if role.lower() not in detail_text:
                    failures.append(
                        f"Labor Benchmarking row {row_number} lists role '{role}' that appears in no "
                        "Milestone Detail row; recost benchmarks must cover only roles priced in the package"
                    )

    element_columns: dict[str, int] = {}
    element_header_row = 0
    for row in detail.iter_rows():
        columns = _header_map(row)
        if "cost element" in columns:
            element_columns = columns
            element_header_row = row[0].row
            break
    if element_columns:
        has_hours = "hours" in element_columns
        has_rate = "rate" in element_columns
        for row_number in range(element_header_row + 1, detail.max_row + 1):
            element = detail.cell(row=row_number, column=element_columns["cost element"]).value
            if not isinstance(element, str) or "labor" not in element.lower():
                continue
            hours_value = detail.cell(row=row_number, column=element_columns["hours"]).value if has_hours else None
            rate_value = detail.cell(row=row_number, column=element_columns["rate"]).value if has_rate else None
            if hours_value is None or rate_value is None:
                failures.append(
                    f"Milestone Detail row {row_number} carries labor delta '{element.strip()}' as a lump sum; "
                    "every recost labor delta must decompose as hours x rate per affected category"
                )

    register_elements = payload.get("recost_register_elements", [])
    if not isinstance(register_elements, list) or not all(
        isinstance(item, str) for item in register_elements
    ):
        raise InputError("recost_register_elements must be an array of strings")
    for element in register_elements:
        if element.strip().lower() not in detail_text:
            failures.append(
                f"change register names cost element '{element.strip()}' but the recost carries no "
                "matching delta row; carry a $0 delta with a one-line justification instead of omitting it"
            )
    return failures


CATEGORY_HEADERS = {"labor category", "category", "role"}
HOURS_DRIVERS = (
    ("fte", "FTE loading"),
    ("duration", "weeks or duration"),
    ("hours_per_period", "hours per FTE-week"),
)
RESTATED_HOURS_EXEMPT_SHEETS = {"Milestone Detail", "Labor Benchmarking"}
RESTATED_HOURS_SKIP = ("productive", "annual", "per year", "/year")


def _normalized_header(value: Any) -> str:
    return re.sub(r"\s+", " ", value).strip().lower() if isinstance(value, str) else ""


def _hours_header_columns(row: Any) -> dict[str, int]:
    """Classify a Milestone Detail labor header row into hours and driver columns."""
    columns: dict[str, int] = {}
    for cell in row:
        header = _normalized_header(cell.value)
        if not header:
            continue
        mentions_hours = "hour" in header or "hrs" in header
        if header in CATEGORY_HEADERS:
            columns.setdefault("category", cell.column)
        elif header == "hours":
            columns.setdefault("hours", cell.column)
        elif mentions_hours and ("week" in header or "fte" in header or "day" in header):
            columns.setdefault("hours_per_period", cell.column)
        elif "fte" in header:
            columns.setdefault("fte", cell.column)
        elif "week" in header or "duration" in header or "month" in header:
            columns.setdefault("duration", cell.column)
    return columns


def _labor_block_rows(sheet: Any, header_row: int, category_column: int) -> list[int]:
    rows: list[int] = []
    for row_number in range(header_row + 1, sheet.max_row + 1):
        label = sheet.cell(row=row_number, column=category_column).value
        if not isinstance(label, str) or not label.strip():
            break
        if "subtotal" in label.lower() or "total" in label.lower():
            break
        rows.append(row_number)
    return rows


def _reconciliation_regions(header_rows: list[int], max_row: int) -> list[tuple[int, int]]:
    regions: list[tuple[int, int]] = []
    for index, header_row in enumerate(header_rows):
        start = 1 if index == 0 else (header_rows[index - 1] + header_row) // 2
        end = max_row if index == len(header_rows) - 1 else (header_row + header_rows[index + 1]) // 2
        regions.append((start, end))
    return regions


def _restated_hours_failures(workbook: Any) -> list[str]:
    failures: list[str] = []
    for sheet_name in REQUIRED_SHEETS:
        if sheet_name in RESTATED_HOURS_EXEMPT_SHEETS or sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                header = _normalized_header(cell.value)
                if "hour" not in header or any(token in header for token in RESTATED_HOURS_SKIP):
                    continue
                seen_value = False
                for row_number in range(cell.row + 1, sheet.max_row + 1):
                    below = sheet.cell(row=row_number, column=cell.column)
                    value = below.value
                    if value is None:
                        if seen_value:
                            break
                        continue
                    seen_value = True
                    if isinstance(value, bool) or not isinstance(value, (int, float)):
                        continue
                    failures.append(
                        f"{sheet_name}!{below.coordinate} restates milestone hours as the constant "
                        f"{float(value):g} under '{str(cell.value).strip()}'; a sheet that repeats Milestone "
                        "Detail hours must reference those cells by formula"
                    )
                    break
    return failures


def hours_derivation_audit(workbook: Any) -> list[str]:
    """Milestone labor hours must be derived from driver input cells, not asserted."""
    failures: list[str] = []
    if "Milestone Detail" not in workbook.sheetnames:
        return failures
    detail = workbook["Milestone Detail"]

    blocks: list[tuple[int, dict[str, int]]] = []
    for row in detail.iter_rows():
        columns = _hours_header_columns(row)
        if "hours" in columns and "category" in columns:
            blocks.append((row[0].row, columns))
    if not blocks:
        return failures

    regions = _reconciliation_regions([header_row for header_row, _ in blocks], detail.max_row)
    for (header_row, columns), (region_start, region_end) in zip(blocks, regions):
        data_rows = _labor_block_rows(detail, header_row, columns["category"])
        missing_drivers = [
            label for key, label in HOURS_DRIVERS if key not in columns
        ]
        if missing_drivers and data_rows:
            failures.append(
                f"Milestone Detail labor block at row {header_row} exposes no "
                f"{', '.join(missing_drivers)} input column; every priced labor row must carry its hours "
                "drivers as input cells"
            )

        asserted: list[int] = []
        for row_number in data_rows:
            hours_cell = detail.cell(row=row_number, column=columns["hours"])
            value = hours_cell.value
            if value is None:
                continue
            if not (isinstance(value, str) and value.startswith("=")):
                asserted.append(row_number)
                continue
            normalized = normalize_formula(value)
            for key, label in HOURS_DRIVERS:
                if key not in columns:
                    continue
                driver = detail.cell(row=row_number, column=columns[key])
                if isinstance(driver.value, bool) or not isinstance(driver.value, (int, float)):
                    failures.append(
                        f"Milestone Detail!{driver.coordinate} {label} driver is not a numeric input cell; "
                        "the hours formula has nothing to recompute from"
                    )
                elif normalize_formula(driver.coordinate) not in normalized:
                    failures.append(
                        f"Milestone Detail!{hours_cell.coordinate} hours formula does not reference its "
                        f"{label} driver cell {driver.coordinate}"
                    )
        if asserted:
            first = detail.cell(row=asserted[0], column=columns["hours"])
            category = detail.cell(row=asserted[0], column=columns["category"]).value
            failures.append(
                f"Milestone Detail!{first.coordinate} asserts hours for "
                f"'{str(category).strip()}' as the constant {first.value} "
                f"({len(asserted)} of {len(data_rows)} labor rows in this block); milestone labor hours must be "
                "a formula over the FTE loading, duration, and hours-per-FTE-week input cells"
            )

        has_reconciliation = False
        has_prose_note = False
        for row in detail.iter_rows(min_row=region_start, max_row=region_end):
            for cell in row:
                value = cell.value
                if not isinstance(value, str):
                    continue
                if value.startswith("="):
                    upper = value.upper()
                    if "MISMATCH" in upper and "OK" in upper:
                        has_reconciliation = True
                elif "hours basis" in value.lower():
                    has_prose_note = True
        if not has_reconciliation:
            prose = " The 'Hours basis:' note is prose, not a check." if has_prose_note else ""
            failures.append(
                f"Milestone Detail rows {region_start}-{region_end} carry no formula reconciling derived "
                "hours to milestone duration and staffing; the reconciliation must be a formula rendering an "
                f"OK or MISMATCH state.{prose}"
            )

    failures.extend(_restated_hours_failures(workbook))
    return failures


# --- Rendered-text clipping audit -------------------------------------------
# A text cell overflows into the next cell only when that neighbour is empty.
# When the neighbour is occupied the label is cut off in the printed workbook,
# so every such label must fit its column, wrap, or be merged across the block.

CLIPPING_GLYPH_WIDTHS = {
    " ": 0.45,
    ".": 0.45,
    ",": 0.45,
    ";": 0.45,
    ":": 0.45,
    "'": 0.35,
    "`": 0.45,
    "!": 0.45,
    "|": 0.45,
    "(": 0.55,
    ")": 0.55,
    "[": 0.55,
    "]": 0.55,
    "{": 0.60,
    "}": 0.60,
    "-": 0.60,
    "/": 0.55,
    "\\": 0.55,
    '"': 0.60,
    "%": 1.50,
    "@": 1.70,
    "$": 1.00,
}
CLIPPING_LOWER_NARROW = "ijl"
CLIPPING_LOWER_SEMI = "frt"
CLIPPING_LOWER_WIDE = "mw"
CLIPPING_UPPER_NARROW = "I"
CLIPPING_UPPER_WIDE = "MW"
CLIPPING_BOLD_FACTOR = 1.14
CLIPPING_ABSOLUTE_TOLERANCE = 0.75
CLIPPING_RELATIVE_TOLERANCE = 0.04
CLIPPING_DEFAULT_WIDTH = 8.43
CLIPPING_MESSAGE_TEXT_LIMIT = 120


def glyph_width(character: str) -> float:
    """Width of one glyph in Excel column-width units (1.0 = one digit)."""
    if character in CLIPPING_GLYPH_WIDTHS:
        return CLIPPING_GLYPH_WIDTHS[character]
    if character in CLIPPING_LOWER_NARROW:
        return 0.48
    if character in CLIPPING_LOWER_SEMI:
        return 0.59
    if character in CLIPPING_LOWER_WIDE:
        return 1.66
    if character in CLIPPING_UPPER_NARROW:
        return 0.45
    if character in CLIPPING_UPPER_WIDE:
        return 1.55
    if character.islower():
        return 0.96
    if character.isupper():
        return 1.05
    return 1.0


def estimated_text_width(text: str, font: Any = None) -> float:
    """Estimated rendered width of a label in column-width units."""
    lines = str(text).split("\n")
    units = max((sum(glyph_width(character) for character in line) for line in lines), default=0.0)
    size = getattr(font, "size", None) or 11.0
    if float(size) != 11.0:
        units *= float(size) / 11.0
    if getattr(font, "bold", False):
        units *= CLIPPING_BOLD_FACTOR
    return units


def column_width_map(sheet: Any) -> tuple[dict[int, float], float]:
    """Explicit column widths by index plus the sheet default width."""
    widths: dict[int, float] = {}
    for letter, dimension in sheet.column_dimensions.items():
        if dimension.width is None:
            continue
        # In-memory dimensions created by a generator carry no min/max, so fall
        # back to the column the dimension is keyed under.
        try:
            own = column_index_from_string(letter)
        except ValueError:
            own = None
        first = dimension.min or own or 1
        last = dimension.max or own or first
        for index in range(first, last + 1):
            widths[index] = float(dimension.width)
    default = sheet.sheet_format.defaultColWidth or CLIPPING_DEFAULT_WIDTH
    return widths, float(default)


def merged_ranges_by_anchor(sheet: Any) -> tuple[dict[tuple[int, int], Any], set[tuple[int, int]]]:
    """Merged-range lookup keyed by anchor cell, plus every covered cell."""
    anchors: dict[tuple[int, int], Any] = {}
    covered: set[tuple[int, int]] = set()
    for merged in sheet.merged_cells.ranges:
        anchors[(merged.min_row, merged.min_col)] = merged
        for row in range(merged.min_row, merged.max_row + 1):
            for column in range(merged.min_col, merged.max_col + 1):
                covered.add((row, column))
    return anchors, covered


def _is_blank(value: Any) -> bool:
    if value is None:
        return True
    return isinstance(value, str) and not value.strip()


def text_clipping_audit(
    workbook: Any,
    *,
    sheets: list[str] | None = None,
    exempt: set[str] | None = None,
) -> list[str]:
    """Flag text that is cut off in print because an occupied neighbour blocks overflow."""
    failures: list[str] = []
    skipped = exempt or set()
    for sheet in workbook.worksheets:
        if sheets is not None and sheet.title not in sheets:
            continue
        widths, default_width = column_width_map(sheet)
        anchors, covered = merged_ranges_by_anchor(sheet)
        max_column = sheet.max_column
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.strip() or value.startswith("="):
                    continue
                position = (cell.row, cell.column)
                if position in covered and position not in anchors:
                    continue
                alignment = cell.alignment
                if alignment.wrap_text or alignment.horizontal in {"fill", "distributed"}:
                    continue
                if f"{sheet.title}!{cell.coordinate}" in skipped:
                    continue
                merged = anchors.get(position)
                first_column = cell.column
                last_column = merged.max_col if merged is not None else cell.column
                available = sum(
                    widths.get(index, default_width)
                    for index in range(first_column, last_column + 1)
                )
                needed = estimated_text_width(value, cell.font)
                tolerance = max(
                    CLIPPING_ABSOLUTE_TOLERANCE,
                    CLIPPING_RELATIVE_TOLERANCE * available,
                )
                if needed <= available + tolerance:
                    continue
                blockers = []
                if alignment.horizontal in {"right", "center", "centerContinuous"}:
                    left = first_column - 1
                    if left < 1:
                        blockers.append("the left sheet edge")
                    elif not _is_blank(sheet.cell(row=cell.row, column=left).value):
                        blockers.append(sheet.cell(row=cell.row, column=left).coordinate)
                if alignment.horizontal != "right":
                    right = last_column + 1
                    if right <= max_column and not _is_blank(
                        sheet.cell(row=cell.row, column=right).value
                    ):
                        blockers.append(sheet.cell(row=cell.row, column=right).coordinate)
                if not blockers:
                    continue
                shown = value.strip()
                if len(shown) > CLIPPING_MESSAGE_TEXT_LIMIT:
                    shown = shown[: CLIPPING_MESSAGE_TEXT_LIMIT - 3] + "..."
                span = (
                    cell.column_letter
                    if merged is None
                    else f"{cell.column_letter}:{get_column_letter(last_column)}"
                )
                failures.append(
                    f"{sheet.title}!{cell.coordinate} is clipped in print: '{shown}' needs about "
                    f"{needed:.1f} column-width units but column {span} gives {available:g} and "
                    f"{', '.join(blockers)} blocks the overflow; widen the column to at least "
                    f"{math.ceil(needed):g}, enable wrap text with adequate row height, merge the "
                    f"label across the block, or shorten it"
                )
    return failures


def structural_audit(workbook: Any, payload: dict[str, Any]) -> list[str]:
    failures: list[str] = []
    for sheet_name in payload.get("required_sheets", REQUIRED_SHEETS):
        if sheet_name not in workbook.sheetnames:
            failures.append(f"missing required sheet: {sheet_name}")

    if "OT Cost Summary" in workbook.sheetnames:
        check_formula(
            failures,
            workbook,
            "'OT Cost Summary'!B8",
            contains=["VALUE(LEFT(B7,4))", "VALUE(LEFT(B6,4))", "VALUE(MID(B7,6,2))", "VALUE(MID(B6,6,2))"],
            not_contains=["DATEDIF", "YEAR("],
        )
        check_formula(
            failures,
            workbook,
            "'OT Cost Summary'!B9",
            contains=["B3", "B8", "^"],
        )
        summary = workbook["OT Cost Summary"]
        if summary["B9"].number_format != "0.0000":
            failures.append("OT Cost Summary!B9 must display the aging factor as 0.0000")

    formula_count = 0
    all_text: list[str] = []
    error_tokens = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str):
                    all_text.append(value)
                    if value.startswith("="):
                        formula_count += 1
                        if any(token in value.upper() for token in error_tokens):
                            failures.append(f"{sheet.title}!{cell.coordinate} contains a formula error token")
                    elif value[:1] in {"+", "-", "@"}:
                        failures.append(f"{sheet.title}!{cell.coordinate} starts with a formula-trigger character")
    if formula_count == 0:
        failures.append("workbook contains no formulas")

    joined = "\n".join(all_text)
    if re.search(r"4022\s*\(d\)\s*\(1\)\s*\(D\).{0,80}competition commitment", joined, re.I | re.S):
        failures.append("workbook misstates 4022(d)(1)(D) as competition commitment")
    if re.search(r"4021.{0,100}(?:100%|fully)\s+Government funded", joined, re.I | re.S):
        failures.append("workbook automatically states 4021 is fully Government funded")
    if re.search(r"4022\s*\(f\).{0,100}(?:100%|fully)\s+Government funded", joined, re.I | re.S):
        failures.append("workbook automatically states 4022(f) is fully Government funded")

    failures.extend(labor_benchmark_audit(workbook))
    failures.extend(hours_reconciliation_audit(workbook))
    failures.extend(hours_derivation_audit(workbook))
    narrative_reported: set[str] = set()
    failures.extend(narrative_format_audit(workbook, narrative_reported))
    failures.extend(text_clipping_audit(workbook, exempt=narrative_reported))
    failures.extend(print_setup_audit(workbook))
    if is_recost_workbook(workbook):
        failures.extend(recost_audit(workbook, payload))

    assertions = payload.get("formula_assertions", [])
    if not isinstance(assertions, list):
        raise InputError("formula_assertions must be an array")
    for index, assertion in enumerate(assertions):
        if not isinstance(assertion, dict) or not isinstance(assertion.get("cell"), str):
            raise InputError(f"formula_assertions[{index}] must contain a cell string")
        contains = assertion.get("contains", [])
        not_contains = assertion.get("not_contains", [])
        if not isinstance(contains, list) or not all(isinstance(item, str) for item in contains):
            raise InputError(f"formula_assertions[{index}].contains must be strings")
        if not isinstance(not_contains, list) or not all(isinstance(item, str) for item in not_contains):
            raise InputError(f"formula_assertions[{index}].not_contains must be strings")
        expected = assertion.get("equals")
        if expected is not None and not isinstance(expected, str):
            raise InputError(f"formula_assertions[{index}].equals must be a string")
        check_formula(
            failures,
            workbook,
            assertion["cell"],
            expected=expected,
            contains=contains,
            not_contains=not_contains,
        )
    return failures


def find_soffice() -> Path | None:
    command = shutil.which("soffice")
    if command:
        return Path(command)
    mac = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    return mac if mac.is_file() else None


def recalculate(source: Path, executable: Path) -> tuple[tempfile.TemporaryDirectory[str], Path]:
    temporary = tempfile.TemporaryDirectory(prefix="ot-workbook-validation-")
    root = Path(temporary.name)
    input_dir = root / "input"
    output_dir = root / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    copied = input_dir / source.name
    shutil.copy2(source, copied)
    completed = subprocess.run(
        [str(executable), "--headless", "--convert-to", "xlsx", "--outdir", str(output_dir), str(copied)],
        capture_output=True,
        text=True,
        timeout=120,
        check=False,
    )
    result = output_dir / source.name
    if completed.returncode != 0 or not result.is_file():
        temporary.cleanup()
        detail = (completed.stderr or completed.stdout).strip()
        raise InputError(f"LibreOffice calculation failed: {detail or 'no output file'}")
    return temporary, result


def close_enough(actual: float, expected: float, tolerance: float) -> bool:
    return math.isclose(actual, expected, rel_tol=tolerance, abs_tol=max(0.01, tolerance))


def compare(workbook: Any, expected: dict[str, Any], tolerance: float) -> list[str]:
    failures: list[str] = []

    def one(reference: str, target: float, label: str) -> None:
        try:
            value = value_at(workbook, reference)
        except InputError as exc:
            failures.append(str(exc))
            return
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            failures.append(f"{label} at {reference} has no calculated numeric value")
        elif not close_enough(float(value), target, tolerance):
            failures.append(f"{label} at {reference} is {float(value):.6f}, expected {target:.6f}")

    for milestone in expected["milestones"]:
        mapping = {
            "workbook_project_cost_cell": ("project_cost", "project cost"),
            "workbook_government_funding_cell": ("government_funding", "Government funding"),
            "workbook_ceiling_cell": ("ceiling_basis", "ceiling basis"),
            "workbook_performer_share_cell": ("performer_project_share", "performer share"),
        }
        for reference_key, (value_key, label) in mapping.items():
            if reference_key in milestone:
                one(milestone[reference_key], milestone[value_key], f"{milestone['id']} {label}")
    if "workbook_total_project_cost_cell" in expected:
        one(expected["workbook_total_project_cost_cell"], expected["total_project_cost"], "total project cost")
    if "workbook_total_government_funding_cell" in expected:
        one(
            expected["workbook_total_government_funding_cell"],
            expected["total_government_funding"],
            "total Government funding",
        )
    return failures


def cached_error_audit(workbook: Any) -> list[str]:
    failures: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("#"):
                    failures.append(f"{sheet.title}!{cell.coordinate} has cached error {value}")
    return failures


def run(path: Path, expected_path: Path, engine: str, tolerance: float) -> dict[str, Any]:
    if not zipfile.is_zipfile(path):
        return {"status": "fail", "failures": ["file is not a valid XLSX ZIP"]}
    payload = load_payload(expected_path)
    expected = calculate(payload)
    workbook = load_workbook(path, data_only=False)
    failures = structural_audit(workbook, payload)
    engine_used = "none"
    temporary: tempfile.TemporaryDirectory[str] | None = None
    try:
        executable = find_soffice() if engine in {"auto", "libreoffice"} else None
        if engine == "libreoffice" and executable is None:
            failures.append("LibreOffice was required but no executable was found")
        elif executable is not None:
            try:
                temporary, calculated_path = recalculate(path, executable)
                calculated = load_workbook(calculated_path, data_only=True)
                failures.extend(cached_error_audit(calculated))
                failures.extend(compare(calculated, expected, tolerance))
                engine_used = "libreoffice"
            except InputError as exc:
                failures.append(str(exc))
    finally:
        if temporary is not None:
            temporary.cleanup()
    return {
        "status": "pass" if not failures else "fail",
        "engine": engine_used,
        "formula_count": sum(
            1
            for sheet in workbook.worksheets
            for row in sheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        ),
        "failures": failures,
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Validate an OT Cost Analysis workbook.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--expected", required=True, type=Path)
    parser.add_argument("--engine", choices=("none", "auto", "libreoffice"), default="auto")
    parser.add_argument("--tolerance", type=float, default=0.01)
    parser.add_argument("--json", action="store_true")
    args = parser.parse_args()
    if not args.workbook.is_file() or not args.expected.is_file():
        print("ERROR: workbook or expected-input file not found", file=sys.stderr)
        return 2
    try:
        result = run(args.workbook, args.expected, args.engine, args.tolerance)
    except (InputError, OSError, ValueError, zipfile.BadZipFile) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2
    if args.json:
        print(json.dumps(result, indent=2, sort_keys=True))
    elif result["status"] == "pass":
        print(f"OT workbook validation passed; formula engine: {result['engine']}.")
    else:
        print("VALIDATION FAILED")
        for failure in result["failures"]:
            print(f"- {failure}")
    return 0 if result["status"] == "pass" else 1


if __name__ == "__main__":
    raise SystemExit(main())
