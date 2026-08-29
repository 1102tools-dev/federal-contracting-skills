#!/usr/bin/env python3
"""Validate an LH/T&M workbook structurally and with optional engine execution."""

from __future__ import annotations

import argparse
import json
import math
import re
import shutil
import subprocess
import sys
import tempfile
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from recompute_expected_values import InputError, calculate, load_payload


DEFAULT_SHEETS = [
    "IGCE Summary",
    "Scenario Analysis",
    "Rate Validation",
    "Travel Detail",
    "Materials Detail",
    "Methodology",
    "Raw Data",
]

CELL_REF = re.compile(
    r"^(?:'((?:[^']|'')+)'|([^!]+))!\$?([A-Za-z]{1,3})\$?([1-9][0-9]*)$"
)
SCENARIO_B_REF = re.compile(
    r"(?:'Scenario Analysis'|Scenario Analysis)!\$?B\$?([1-9][0-9]*)",
    re.IGNORECASE,
)
BASE_PERIOD_LABEL = re.compile(r"\bBASE[\s-]*(?:YEAR|PERIOD)\b")
OPTION_PERIOD_LABEL = re.compile(
    r"\bOPTION\s*(?:YEAR|PERIOD)?\s*0*([1-9][0-9]*)\b|\bO[YP]\s*0*([1-9][0-9]*)\b"
)
TOTAL_PERIODS_CELL = "B14"
SUMMARY_CONSTANT_THRESHOLD = 1000.0
MONEY_KEYWORDS = (
    "TOTAL",
    "COST",
    "PRICE",
    "AMOUNT",
    "ODC",
    "OTHER DIRECT",
    "TRAVEL",
    "MATERIALS",
    "ESTIMATE",
)
NON_MONEY_KEYWORDS = ("HOUR", "FTE", "MONTH", "SOC", "PERIODS")
PERIOD_TOTAL_LABEL = re.compile(
    r"\bTOTAL\b.*\bALL\s+(?:PERIODS|YEARS)\b|\bALL\s+(?:PERIODS|YEARS)\b.*\bTOTAL\b"
)
SCENARIO_MONEY_KEYWORDS = MONEY_KEYWORDS + ("LABOR",)
SCENARIO_NON_MONEY_HEADERS = (
    "FACTOR",
    "MULTIPLIER",
    "RATE",
    "HOUR",
    "FTE",
    "ESCALATION",
    "PERCENT",
)
SUMMARY_PERIOD_SKIP = ("MONTH", "HOUR", "FTE", "SOC")
PERIOD_TIE_TOLERANCE = 1.0


def normalize_formula(value: Any) -> str:
    return re.sub(r"\s+", "", str(value)).upper()


def parse_cell_ref(reference: str) -> tuple[str, str]:
    match = CELL_REF.fullmatch(reference.strip())
    if not match:
        raise InputError(f"invalid workbook cell reference: {reference}")
    sheet = (match.group(1) or match.group(2)).replace("''", "'")
    coordinate = f"{match.group(3).upper()}{match.group(4)}"
    return sheet, coordinate


def cell_value(workbook: Any, reference: str) -> Any:
    sheet, coordinate = parse_cell_ref(reference)
    if sheet not in workbook.sheetnames:
        raise InputError(f"cell reference uses missing sheet: {reference}")
    return workbook[sheet][coordinate].value


def check_formula(
    failures: list[str],
    sheet: Any,
    coordinate: str,
    *,
    expected: str | None = None,
    contains: list[str] | None = None,
    not_contains: list[str] | None = None,
) -> None:
    value = sheet[coordinate].value
    if not isinstance(value, str) or not value.startswith("="):
        failures.append(f"{sheet.title}!{coordinate} is not a formula")
        return
    normalized = normalize_formula(value)
    if expected is not None and normalized != normalize_formula(expected):
        failures.append(f"{sheet.title}!{coordinate} formula does not match expected structure")
    for item in contains or []:
        if normalize_formula(item) not in normalized:
            failures.append(f"{sheet.title}!{coordinate} formula is missing {item}")
    for item in not_contains or []:
        if normalize_formula(item) in normalized:
            failures.append(f"{sheet.title}!{coordinate} formula contains forbidden text {item}")


def _is_number(value: Any) -> bool:
    return not isinstance(value, bool) and isinstance(value, (int, float))


def _nearest_header(sheet: Any, row: int, column: int, limit: int = 40) -> str:
    """Return the closest non-formula text above a cell in the same column."""
    for row_index in range(row - 1, max(0, row - limit), -1):
        value = sheet.cell(row_index, column).value
        if isinstance(value, str) and value.strip() and not value.startswith("="):
            return value.upper()
    return ""


def effective_periods(workbook: Any, payload_periods: int, failures: list[str]) -> int:
    """Reconcile the Total Periods assumption cell with the validation input."""
    if "IGCE Summary" not in workbook.sheetnames:
        return payload_periods
    value = workbook["IGCE Summary"][TOTAL_PERIODS_CELL].value
    if value is None:
        if payload_periods > 1:
            failures.append(
                f"IGCE Summary!{TOTAL_PERIODS_CELL} must record Total Periods (Base plus "
                f"Options) for a {payload_periods}-period requirement"
            )
        return payload_periods
    if not _is_number(value) or float(value) != int(value) or int(value) < 1:
        failures.append(
            f"IGCE Summary!{TOTAL_PERIODS_CELL} Total Periods (Base plus Options) must be "
            "a whole number of at least 1"
        )
        return payload_periods
    declared = int(value)
    if payload_periods > 1 and declared != payload_periods:
        failures.append(
            f"IGCE Summary!{TOTAL_PERIODS_CELL} Total Periods is {declared} but the "
            f"validation input states {payload_periods} periods"
        )
    return max(declared, payload_periods)


def period_coverage_audit(workbook: Any, periods: int) -> list[str]:
    """Require per-period totals on IGCE Summary for a multi-period requirement."""
    failures: list[str] = []
    if periods <= 1 or "IGCE Summary" not in workbook.sheetnames:
        return failures
    base_found = False
    option_numbers: set[int] = set()
    for row in workbook["IGCE Summary"].iter_rows():
        for cell in row:
            value = cell.value
            if not isinstance(value, str):
                continue
            upper = value.upper()
            if BASE_PERIOD_LABEL.search(upper):
                base_found = True
            for match in OPTION_PERIOD_LABEL.finditer(upper):
                option_numbers.add(int(match.group(1) or match.group(2)))
    if not base_found or len(option_numbers) < periods - 1:
        found = ", ".join(f"Option {n}" for n in sorted(option_numbers)) or "none"
        failures.append(
            f"the stated period of performance has {periods} periods (base plus "
            f"{periods - 1} options) but IGCE Summary shows no per-period totals (base "
            f"label found: {'yes' if base_found else 'no'}; option-period labels found: "
            f"{found}); a headline planning estimate must cover every period or carry an "
            "explicit single-period coverage annotation for a single-period scope"
        )
    return failures


def escalation_input_audit(workbook: Any, periods: int) -> list[str]:
    """Reject a dead escalation input when multi-period pricing exists."""
    failures: list[str] = []
    if periods <= 1 or "IGCE Summary" not in workbook.sheetnames:
        return failures
    summary = workbook["IGCE Summary"]
    target = None
    for row in summary.iter_rows():
        for cell in row:
            value = cell.value
            if (
                isinstance(value, str)
                and "ESCALATION" in value.upper()
                and len(value.strip()) <= 30
            ):
                for column in range(cell.column + 1, min(cell.column + 6, summary.max_column) + 1):
                    candidate = summary.cell(cell.row, column)
                    if _is_number(candidate.value):
                        target = candidate
                        break
                if target is None and _is_number(summary.cell(cell.row + 1, cell.column).value):
                    target = summary.cell(cell.row + 1, cell.column)
            if target is not None:
                break
        if target is not None:
            break
    if target is None:
        failures.append(
            "IGCE Summary has no numeric Escalation Rate input; a multi-period estimate "
            "must carry a live escalation input"
        )
        return failures
    column_letter = target.column_letter
    row_number = target.row
    bare_reference = re.compile(
        rf"(?<![A-Z0-9_$])\$?{column_letter}\$?{row_number}(?![0-9])"
    )
    sheet_reference = re.compile(
        rf"(?:'IGCE SUMMARY'|IGCE SUMMARY)!\$?{column_letter}\$?{row_number}(?![0-9])"
    )
    referenced = False
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                normalized = normalize_formula(value)
                if sheet.title == "IGCE Summary":
                    if bare_reference.search(normalized):
                        referenced = True
                elif sheet_reference.search(normalized):
                    referenced = True
            if referenced:
                break
        if referenced:
            break
    if not referenced:
        failures.append(
            f"IGCE Summary!{target.coordinate} escalation input "
            f"({target.value}) is referenced by zero formulas; multi-period pricing must "
            "apply the escalation input rather than displaying a dead cell"
        )
    return failures


def summary_constant_audit(workbook: Any) -> list[str]:
    """Reject hardcoded summary money amounts absent from the refresh register."""
    failures: list[str] = []
    if "IGCE Summary" not in workbook.sheetnames or "Raw Data" not in workbook.sheetnames:
        return failures
    register_numbers: list[float] = []
    for row in workbook["Raw Data"].iter_rows():
        for cell in row:
            value = cell.value
            if _is_number(value):
                register_numbers.append(float(value))
            elif isinstance(value, str):
                for token in re.findall(r"\d[\d,]*(?:\.\d+)?", value):
                    register_numbers.append(float(token.replace(",", "")))
    summary = workbook["IGCE Summary"]
    for row in summary.iter_rows():
        label = next(
            (c.value for c in row if isinstance(c.value, str) and c.value.strip()), ""
        )
        label_upper = str(label).upper()
        for cell in row:
            value = cell.value
            if not _is_number(value) or abs(float(value)) < SUMMARY_CONSTANT_THRESHOLD:
                continue
            header = _nearest_header(summary, cell.row, cell.column)
            if any(keyword in header for keyword in NON_MONEY_KEYWORDS):
                continue
            if not any(
                keyword in header or keyword in label_upper for keyword in MONEY_KEYWORDS
            ):
                continue
            if any(abs(float(value) - entry) <= 1.0 for entry in register_numbers):
                continue
            failures.append(
                f"IGCE Summary!{cell.coordinate} hardcodes {float(value):,.2f} "
                f"({str(label).strip() or 'unlabeled line'}) with no matching entry in "
                "the Raw Data refresh register; every summary money line must trace to a "
                "detail sheet or assumptions input and appear in the register"
            )
    return failures


def _period_key(label: str) -> tuple[Any, ...] | None:
    """Classify a row label as a period key: total, option N, or base."""
    if PERIOD_TOTAL_LABEL.search(label):
        return ("total",)
    match = OPTION_PERIOD_LABEL.search(label)
    if match:
        return ("option", int(match.group(1) or match.group(2)))
    if BASE_PERIOD_LABEL.search(label):
        return ("base",)
    return None


def _period_name(key: tuple[Any, ...]) -> str:
    if key[0] == "total":
        return "the all-periods total"
    if key[0] == "option":
        return f"option period {key[1]}"
    return "the base period"


def _row_label(row: tuple[Any, ...]) -> str:
    return next(
        (c.value for c in row if isinstance(c.value, str) and c.value.strip()), ""
    ).upper()


def scenario_period_tie_audit(values_workbook: Any) -> list[str]:
    """Require Scenario Analysis period money figures to tie to IGCE Summary.

    Runs on calculated (data_only) values, so a stale formula whose text looks
    plausible but whose result is zero is still caught. Money-labeled scenario
    period cells that are all zero while the summary per-period totals are
    nonzero fail outright; otherwise every current-assumptions (mid) scenario
    figure must agree with an IGCE Summary amount for the same period within $1.
    Cells with no calculated value available are skipped.
    """
    failures: list[str] = []
    if (
        "Scenario Analysis" not in values_workbook.sheetnames
        or "IGCE Summary" not in values_workbook.sheetnames
    ):
        return failures

    summary_periods: dict[tuple[Any, ...], list[tuple[str, float]]] = {}
    for row in values_workbook["IGCE Summary"].iter_rows():
        label = _row_label(row)
        key = _period_key(label)
        if key is None or any(word in label for word in SUMMARY_PERIOD_SKIP):
            continue
        for cell in row:
            if _is_number(cell.value):
                summary_periods.setdefault(key, []).append(
                    (cell.coordinate, float(cell.value))
                )

    scenario = values_workbook["Scenario Analysis"]
    money_cells: list[tuple[tuple[Any, ...], Any, str]] = []
    for row in scenario.iter_rows():
        key = _period_key(_row_label(row))
        if key is None:
            continue
        for cell in row:
            if not _is_number(cell.value):
                continue
            header = _nearest_header(scenario, cell.row, cell.column, limit=10)
            if any(word in header for word in SCENARIO_NON_MONEY_HEADERS):
                continue
            if not any(word in header for word in SCENARIO_MONEY_KEYWORDS):
                continue
            money_cells.append((key, cell, header))
    if not money_cells:
        return failures

    summary_nonzero = any(
        abs(amount) > PERIOD_TIE_TOLERANCE
        for entries in summary_periods.values()
        for _, amount in entries
    )
    if summary_nonzero and all(
        abs(float(cell.value)) < 0.005 for _, cell, _ in money_cells
    ):
        coordinates = ", ".join(cell.coordinate for _, cell, _ in money_cells)
        failures.append(
            f"Scenario Analysis period-totals money cells ({coordinates}) are all zero "
            "while IGCE Summary per-period totals are nonzero; the scenario table is "
            "disconnected from the live summary figures (check the formulas for stale "
            "ranges that point at the wrong rows)"
        )
        return failures

    has_mid = any("MID" in h or "POINT" in h for _, _, h in money_cells)
    for key, cell, header in money_cells:
        if has_mid and "MID" not in header and "POINT" not in header:
            continue
        entries = summary_periods.get(key)
        if not entries:
            continue
        value = float(cell.value)
        if not any(
            abs(value - amount) <= PERIOD_TIE_TOLERANCE for _, amount in entries
        ):
            closest = min(entries, key=lambda item: abs(value - item[1]))
            failures.append(
                f"Scenario Analysis!{cell.coordinate} shows {value:,.2f} for "
                f"{_period_name(key)} under current assumptions but no IGCE Summary "
                f"amount for that period is within $1 (closest is IGCE Summary!"
                f"{closest[0]} at {closest[1]:,.2f}); the scenario period totals must "
                "tie to the IGCE Summary per-period totals"
            )
    return failures


def material_handling_audit(workbook: Any, payload: dict[str, Any]) -> list[str]:
    """Reject undisclosed material-handling amounts or formulas.

    Zero remains the default. A nonzero numeric input or formula is allowed only
    when the validation sidecar identifies the exact cell and value/formula and
    records the user-supplied accounting or solicitation basis.
    """

    failures: list[str] = []
    if "Materials Detail" not in workbook.sheetnames:
        return failures

    assumptions = payload.get("assumptions", {})
    contract_type = assumptions.get("contract_type", "") if isinstance(assumptions, dict) else ""
    is_tm = isinstance(contract_type, str) and contract_type.upper() in {"T&M", "TM"}
    sheet = workbook["Materials Detail"]

    headers: list[Any] = []
    for row in sheet.iter_rows():
        for cell in row:
            label = re.sub(r"[^A-Z0-9]+", " ", str(cell.value or "").upper()).strip()
            if label in {
                "MATERIAL HANDLING",
                "MATERIAL HANDLING COST",
                "MATERIAL HANDLING INDIRECT",
                "MATERIAL HANDLING INDIRECT COST",
                "APPLICABLE MATERIAL HANDLING INDIRECT COST",
            }:
                headers.append(cell)
    if is_tm and not headers:
        failures.append("Materials Detail has no Material Handling column")
        return failures

    raw_assertions = payload.get("material_handling_assertions", [])
    if not isinstance(raw_assertions, list):
        raise InputError("material_handling_assertions must be an array")

    assertions: dict[str, dict[str, Any]] = {}
    for index, assertion in enumerate(raw_assertions):
        if not isinstance(assertion, dict):
            raise InputError(f"material_handling_assertions[{index}] must be an object")
        reference = assertion.get("cell")
        basis = assertion.get("basis")
        expected = assertion.get("equals")
        if not isinstance(reference, str):
            raise InputError(f"material_handling_assertions[{index}].cell must be a string")
        if not isinstance(basis, str) or not basis.strip():
            raise InputError(
                f"material_handling_assertions[{index}].basis must be a non-empty string"
            )
        if isinstance(expected, bool) or not isinstance(expected, (int, float, str)):
            raise InputError(
                f"material_handling_assertions[{index}].equals must be a number or formula string"
            )
        if isinstance(expected, (int, float)) and (
            not math.isfinite(float(expected)) or float(expected) < 0
        ):
            raise InputError(
                f"material_handling_assertions[{index}].equals must be finite and non-negative"
            )
        if isinstance(expected, str) and not expected.startswith("="):
            raise InputError(
                f"material_handling_assertions[{index}].equals string must be a formula"
            )
        sheet_name, coordinate = parse_cell_ref(reference)
        if sheet_name != "Materials Detail":
            raise InputError(
                f"material_handling_assertions[{index}].cell must reference Materials Detail"
            )
        normalized_reference = f"Materials Detail!{coordinate}"
        if normalized_reference in assertions:
            raise InputError(f"duplicate material-handling assertion: {reference}")
        assertions[normalized_reference] = {
            "equals": expected,
            "basis": basis.strip(),
        }

    inspected: set[str] = set()
    for header in headers:
        for row_number in range(header.row + 1, sheet.max_row + 1):
            cell = sheet.cell(row_number, header.column)
            value = cell.value
            if value is None or value == "":
                continue
            reference = f"Materials Detail!{cell.coordinate}"
            inspected.add(reference)
            assertion = assertions.get(reference)

            if isinstance(value, bool):
                failures.append(f"{reference} must be a numeric input or disclosed formula")
                continue
            if isinstance(value, (int, float)) and float(value) == 0 and assertion is None:
                continue
            if not is_tm:
                failures.append(f"{reference} must be zero for a Labor-Hour estimate")
                continue
            if assertion is None:
                failures.append(
                    f"{reference} contains undisclosed material handling {value!r}; "
                    "zero is required unless the sidecar records the exact value/formula and basis"
                )
                continue

            expected = assertion["equals"]
            if isinstance(expected, str):
                if not isinstance(value, str) or normalize_formula(value) != normalize_formula(expected):
                    failures.append(f"{reference} does not match its disclosed material-handling formula")
            elif isinstance(value, bool) or not isinstance(value, (int, float)):
                failures.append(f"{reference} is not the disclosed numeric material-handling input")
            elif not math.isclose(float(value), float(expected), rel_tol=0, abs_tol=1e-9):
                failures.append(
                    f"{reference} is {float(value):.6f}, expected disclosed input {float(expected):.6f}"
                )

    for reference in sorted(set(assertions) - inspected):
        failures.append(f"material-handling assertion references no populated handling cell: {reference}")
    return failures


def structural_audit(workbook: Any, payload: dict[str, Any]) -> list[str]:
    failures: list[str] = []
    required_sheets = payload.get("required_sheets", DEFAULT_SHEETS)
    if not isinstance(required_sheets, list) or not all(
        isinstance(item, str) for item in required_sheets
    ):
        raise InputError("required_sheets must be an array of strings")
    for sheet_name in required_sheets:
        if sheet_name not in workbook.sheetnames:
            failures.append(f"missing required sheet: {sheet_name}")

    if "IGCE Summary" in workbook.sheetnames:
        summary = workbook["IGCE Summary"]
        check_formula(
            failures,
            summary,
            "B10",
            contains=[
                "VALUE(LEFT(B9,4))",
                "VALUE(MID(B9,6,2))",
                "VALUE(LEFT(B8,4))",
                "VALUE(MID(B8,6,2))",
            ],
            not_contains=["DATEDIF", "YEAR("],
        )
        check_formula(
            failures,
            summary,
            "B11",
            contains=["B5", "B10", "^"],
        )
        if summary["B11"].number_format != "0.0000":
            failures.append("IGCE Summary!B11 must display the aging factor as 0.0000")

    formula_count = 0
    formula_error_tokens = ("#REF!", "#NAME?", "#VALUE!", "#DIV/0!")
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("="):
                    formula_count += 1
                    upper = value.upper()
                    if any(token in upper for token in formula_error_tokens):
                        failures.append(f"{sheet.title}!{cell.coordinate} contains a formula error token")
                elif isinstance(value, str) and value[:1] in {"+", "-", "@"}:
                    failures.append(
                        f"{sheet.title}!{cell.coordinate} starts with a formula-trigger character"
                    )
    if formula_count == 0:
        failures.append("workbook contains no formulas")

    if "Scenario Analysis" in workbook.sheetnames:
        buildup = workbook["Scenario Analysis"]
        starts: list[int] = []
        for row_index in range(1, buildup.max_row + 1):
            label = buildup.cell(row_index, 1).value
            if isinstance(label, str) and label.startswith("Scenario Analysis:"):
                starts.append(row_index)
        if not starts:
            failures.append("Scenario Analysis contains no recognized labor blocks")
        for block_index, base in enumerate(starts):
            expected_base = 1 + block_index * 13
            if base != expected_base:
                failures.append(
                    f"Scenario Analysis block {block_index + 1} starts at row {base}, expected {expected_base}"
                )
            if buildup[f"B{base + 5}"].value is not None:
                failures.append(f"Scenario Analysis!B{base + 5} must be the blank separator row")
            if buildup[f"B{base + 12}"].value is not None:
                failures.append(f"Scenario Analysis!B{base + 12} must be the blank block separator")
            check_formula(
                failures,
                buildup,
                f"B{base + 2}",
                contains=["'IGCE SUMMARY'!$B$11"],
            )
            if buildup[f"B{base + 2}"].number_format != "0.0000":
                failures.append(
                    f"Scenario Analysis!B{base + 2} must display the aging factor as 0.0000"
                )
            formulas = {
                base + 3: f"=B{base + 1}*B{base + 2}",
                base + 4: f"=B{base + 3}/2080",
                base + 6: "='IGCE Summary'!$B$2",
                base + 7: f"=B{base + 4}*B{base + 6}",
                base + 8: "='IGCE Summary'!$B$3",
                base + 9: f"=B{base + 4}*B{base + 8}",
                base + 10: "='IGCE Summary'!$B$4",
                base + 11: f"=B{base + 4}*B{base + 10}",
            }
            for row_number, expected_formula in formulas.items():
                check_formula(
                    failures,
                    buildup,
                    f"B{row_number}",
                    expected=expected_formula,
                )

    for sheet_name in ("IGCE Summary", "Rate Validation"):
        if sheet_name not in workbook.sheetnames:
            continue
        sheet = workbook[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.startswith("="):
                    continue
                for match in SCENARIO_B_REF.finditer(value):
                    referenced_row = int(match.group(1))
                    if (referenced_row - 4) % 13 == 0:
                        failures.append(
                            f"{sheet_name}!{cell.coordinate} uses Scenario Analysis row {referenced_row} "
                            "as a cross-sheet input; that row is Aged Annual Wage, not Direct Labor"
                        )

    failures.extend(material_handling_audit(workbook, payload))

    assertions = payload.get("formula_assertions", [])
    if not isinstance(assertions, list):
        raise InputError("formula_assertions must be an array")
    for index, assertion in enumerate(assertions):
        if not isinstance(assertion, dict) or not isinstance(assertion.get("cell"), str):
            raise InputError(f"formula_assertions[{index}] must contain a cell string")
        sheet_name, coordinate = parse_cell_ref(assertion["cell"])
        if sheet_name not in workbook.sheetnames:
            failures.append(f"formula assertion references missing sheet: {sheet_name}")
            continue
        contains = assertion.get("contains", [])
        not_contains = assertion.get("not_contains", [])
        if not isinstance(contains, list) or not all(isinstance(item, str) for item in contains):
            raise InputError(f"formula_assertions[{index}].contains must be an array of strings")
        if not isinstance(not_contains, list) or not all(
            isinstance(item, str) for item in not_contains
        ):
            raise InputError(
                f"formula_assertions[{index}].not_contains must be an array of strings"
            )
        equals = assertion.get("equals")
        if equals is not None and not isinstance(equals, str):
            raise InputError(f"formula_assertions[{index}].equals must be a string")
        check_formula(
            failures,
            workbook[sheet_name],
            coordinate,
            expected=equals,
            contains=contains,
            not_contains=not_contains,
        )
    return failures


def find_soffice() -> Path | None:
    discovered = shutil.which("soffice")
    if discovered:
        return Path(discovered)
    mac_path = Path("/Applications/LibreOffice.app/Contents/MacOS/soffice")
    return mac_path if mac_path.is_file() else None


def recalculate_with_libreoffice(source: Path, executable: Path) -> tuple[tempfile.TemporaryDirectory[str], Path]:
    temp = tempfile.TemporaryDirectory(prefix="lh-tm-workbook-validation-")
    temp_root = Path(temp.name)
    input_dir = temp_root / "input"
    output_dir = temp_root / "output"
    input_dir.mkdir()
    output_dir.mkdir()
    copied = input_dir / source.name
    shutil.copy2(source, copied)
    command = [
        str(executable),
        "--headless",
        "--convert-to",
        "xlsx",
        "--outdir",
        str(output_dir),
        str(copied),
    ]
    completed = subprocess.run(command, capture_output=True, text=True, timeout=120, check=False)
    recalculated = output_dir / source.name
    if completed.returncode != 0 or not recalculated.is_file():
        temp.cleanup()
        detail = (completed.stderr or completed.stdout).strip()
        raise InputError(f"LibreOffice recalculation failed: {detail or 'no output file'}")
    return temp, recalculated


def close_enough(actual: float, expected: float, tolerance: float) -> bool:
    return math.isclose(actual, expected, rel_tol=tolerance, abs_tol=max(0.01, tolerance))


def compare_results(
    workbook: Any,
    expected: dict[str, Any],
    tolerance: float,
) -> list[str]:
    failures: list[str] = []

    def compare(reference: str, target: float, label: str) -> None:
        try:
            value = cell_value(workbook, reference)
        except InputError as exc:
            failures.append(str(exc))
            return
        if isinstance(value, bool) or not isinstance(value, (int, float)):
            failures.append(f"{label} at {reference} has no calculated numeric value")
            return
        if not close_enough(float(value), target, tolerance):
            failures.append(
                f"{label} at {reference} is {float(value):.6f}, expected {target:.6f}"
            )

    for line in expected["labor_lines"]:
        mappings = (
            ("workbook_low_rate_cell", "burdened_low_rate", "low rate"),
            ("workbook_mid_rate_cell", "burdened_mid_rate", "mid rate"),
            ("workbook_high_rate_cell", "burdened_high_rate", "high rate"),
            ("workbook_mid_total_cell", "labor_mid_total", "mid labor total"),
        )
        for reference_key, expected_key, label in mappings:
            if reference_key in line:
                compare(line[reference_key], line[expected_key], f"{line['name']} {label}")
    for line in expected["non_labor_lines"]:
        if "workbook_total_cell" in line:
            compare(
                line["workbook_total_cell"],
                line["amount"],
                f"{line['name']} amount",
            )
    totals = (
        ("workbook_low_total_cell", "low_estimated_total", "low estimated total"),
        ("workbook_mid_total_cell", "mid_estimated_total", "mid estimated total"),
        ("workbook_high_total_cell", "high_estimated_total", "high estimated total"),
        ("workbook_ceiling_price_cell", "ceiling_price", "ceiling price"),
    )
    for reference_key, expected_key, label in totals:
        if reference_key in expected and expected_key in expected:
            compare(expected[reference_key], expected[expected_key], label)
    return failures


def cached_error_audit(workbook: Any) -> list[str]:
    failures: list[str] = []
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if isinstance(value, str) and value.startswith("#"):
                    failures.append(f"{sheet.title}!{cell.coordinate} has cached error {value}")
    return failures


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Audit an LH/T&M workbook and optionally verify formula execution with LibreOffice."
    )
    parser.add_argument("workbook", type=Path, help="Workbook to validate")
    parser.add_argument("--expected", required=True, type=Path, help="Validation-input JSON")
    parser.add_argument(
        "--engine",
        choices=("auto", "none", "libreoffice"),
        default="auto",
        help="Formula execution engine policy",
    )
    parser.add_argument("--tolerance", type=float, default=0.01, help="Relative comparison tolerance")
    parser.add_argument("--json", action="store_true", help="Emit a JSON result")
    args = parser.parse_args()

    if not math.isfinite(args.tolerance) or args.tolerance < 0:
        print("ERROR: tolerance must be finite and non-negative", file=sys.stderr)
        return 2
    if not args.workbook.is_file():
        print(f"ERROR: workbook not found: {args.workbook}", file=sys.stderr)
        return 2

    try:
        payload = load_payload(args.expected)
        expected = calculate(payload)
        formula_workbook = load_workbook(args.workbook, data_only=False)
        structural_failures = structural_audit(formula_workbook, payload)
        periods = effective_periods(formula_workbook, expected["periods"], structural_failures)
        structural_failures.extend(period_coverage_audit(formula_workbook, periods))
        structural_failures.extend(escalation_input_audit(formula_workbook, periods))
        structural_failures.extend(summary_constant_audit(formula_workbook))
        failures = list(structural_failures)
        values_workbook = load_workbook(args.workbook, data_only=True)
        failures.extend(scenario_period_tie_audit(values_workbook))
    except (InputError, OSError, ValueError) as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 2

    engine_used: str | None = None
    engine_note: str
    temp: tempfile.TemporaryDirectory[str] | None = None
    try:
        if args.engine == "none":
            engine_note = "Formula execution was not requested."
        else:
            soffice = find_soffice()
            if soffice is None:
                if args.engine == "libreoffice":
                    failures.append("LibreOffice was required but soffice was not found")
                    engine_note = "LibreOffice was required but unavailable."
                else:
                    engine_note = "Formula execution was not independently verified in Excel or LibreOffice."
            else:
                temp, recalculated = recalculate_with_libreoffice(args.workbook, soffice)
                calculated_workbook = load_workbook(recalculated, data_only=True)
                failures.extend(cached_error_audit(calculated_workbook))
                failures.extend(compare_results(calculated_workbook, expected, args.tolerance))
                for failure in scenario_period_tie_audit(calculated_workbook):
                    if failure not in failures:
                        failures.append(failure)
                engine_used = "LibreOffice"
                engine_note = "LibreOffice formula execution and cached-value comparison ran."
    except (InputError, OSError, subprocess.SubprocessError, ValueError) as exc:
        failures.append(str(exc))
        engine_note = "LibreOffice execution failed."
    finally:
        if temp is not None:
            temp.cleanup()

    result = {
        "status": "pass" if not failures else "fail",
        "formula_structure": "pass" if not structural_failures else "fail",
        "independent_recomputation": "pass",
        "independent_mid_estimated_total": expected["mid_estimated_total"],
        "engine": engine_used,
        "engine_note": engine_note,
        "failures": failures,
    }
    if args.json:
        print(json.dumps(result, indent=2, sort_keys=True))
    else:
        if failures:
            print("VALIDATION FAILED")
            for failure in failures:
                print(f"- {failure}")
        elif engine_used:
            print("Formula structure, independent calculations, and LibreOffice formula execution passed.")
        else:
            print(
                "Formula structure and independent calculations passed. "
                "Formula execution was not independently verified in Excel or LibreOffice."
            )
        print(f"Independent mid estimated total: {expected['mid_estimated_total']:.2f}")
        print(engine_note)
    return 0 if not failures else 1


if __name__ == "__main__":
    raise SystemExit(main())
