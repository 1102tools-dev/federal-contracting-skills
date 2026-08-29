from __future__ import annotations

import importlib.util
import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


ROOT = Path(__file__).resolve().parents[1]
VALIDATOR = ROOT / "skills" / "igce-builder-cr" / "scripts" / "validate_workbook.py"


def load_clipping_validator(skill: str):
    directory = ROOT / "skills" / skill / "scripts"
    spec = importlib.util.spec_from_file_location(
        f"{skill.replace('-', '_')}_clipping_tests",
        directory / "validate_workbook.py",
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.path.insert(0, str(directory))
    try:
        spec.loader.exec_module(module)
    finally:
        sys.path.pop(0)
    return module


def autosize_for_print(workbook: Workbook, skill: str) -> Workbook:
    """Set every column width from the longest label actually written to it.

    This is the technique workbook-specification.md now requires of the
    generators: text cannot overflow into an occupied neighbour, so the column
    must be wide enough for its own widest label.
    """
    module = load_clipping_validator(skill)
    for sheet in workbook.worksheets:
        needed: dict[str, float] = {}
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.strip() or value.startswith("="):
                    continue
                width = module.estimated_text_width(value, cell.font)
                letter = cell.column_letter
                needed[letter] = max(needed.get(letter, 0.0), width)
        for letter, width in needed.items():
            current = sheet.column_dimensions[letter].width or 0.0
            sheet.column_dimensions[letter].width = max(current, min(90.0, width + 2.0))
    return workbook


def fixture_workbook() -> Workbook:
    workbook = Workbook()
    workbook.active.title = "IGCE Summary"
    for name in (
        "Cost Buildup",
        "Scenario Analysis",
        "Rate Validation",
        "Travel Detail",
        "Methodology",
        "Raw Data",
    ):
        workbook.create_sheet(name)

    summary = workbook["IGCE Summary"]
    summary["B2"] = 0.32
    summary["B3"] = 0.45
    summary["B4"] = 0.12
    summary["B5"] = 0.005
    summary["B6"] = 0.025
    summary["B7"] = 1880
    summary["B8"] = 12
    summary["B9"] = "2026-01"
    summary["B10"] = "2026-01"
    summary["B11"] = (
        "=MAX(0,(VALUE(LEFT(B10,4))-VALUE(LEFT(B9,4)))*12+"
        "VALUE(MID(B10,6,2))-VALUE(MID(B9,6,2)))"
    )
    summary["B12"] = "=(1+B6)^(B11/12)"
    summary["B12"].number_format = "0.0000"
    summary["B13"] = "CPFF"
    summary["B14"] = 0.07
    summary["A19"] = "Total Periods (Base plus Options)"
    summary["B19"] = 1
    summary["A21"] = "Total estimated cost"
    summary["B21"] = 1000000
    summary["A22"] = "Fee"
    summary["B22"] = 70000
    summary["A23"] = "Total estimated price"
    summary["B23"] = 1070000

    buildup = workbook["Cost Buildup"]
    buildup["A1"] = "Cost Buildup: Test Analyst"
    buildup["B2"] = 100000
    buildup["B3"] = "='IGCE Summary'!$B$12"
    buildup["B3"].number_format = "0.0000"
    buildup["B4"] = "=B2*B3"
    buildup["B5"] = "=B4/2080"
    buildup["B7"] = "='IGCE Summary'!$B$2"
    buildup["B8"] = "=B5*B7"
    buildup["B9"] = "=B5+B8"
    buildup["B10"] = "='IGCE Summary'!$B$3"
    buildup["B11"] = "=B9*B10"
    buildup["B12"] = "=B9+B11"
    buildup["B13"] = "='IGCE Summary'!$B$4"
    buildup["B14"] = "=B12*B13"
    buildup["B15"] = "='IGCE Summary'!$B$5"
    buildup["B16"] = "=(B12+B14)*B15"
    buildup["B17"] = "=B12+B14+B16"
    buildup["B18"] = "='IGCE Summary'!$B$13"
    buildup["B19"] = "='IGCE Summary'!$B$14"
    buildup["B20"] = "=B17*'IGCE Summary'!$B$14"
    buildup["B21"] = "=B17+B20"
    buildup["B22"] = "=B21/B5"

    scenario = workbook["Scenario Analysis"]
    scenario["A1"] = "CPFF Scenario Analysis"
    scenario["A3"] = "Scenario"
    scenario["B3"] = "Estimated cost"
    scenario["C3"] = "Fee"
    scenario["D3"] = "Estimated price"
    scenario["A4"] = "Working"
    scenario["B4"] = 1000000
    scenario["C4"] = 70000
    scenario["D4"] = 1070000
    return autosize_for_print(workbook, "igce-builder-cr")


def fixture_payload() -> dict:
    return {
        "assumptions": {
            "fee_type": "CPFF",
            "primary_fee_rate": 0.07,
            "fringe_rate": 0.32,
            "overhead_rate": 0.45,
            "ga_rate": 0.12,
            "fccm_rate": 0.005,
            "aging_factor": 1.0,
            "productive_hours": 1880,
        },
        "labor_lines": [
            {
                "name": "Test Analyst",
                "annual_wage": 100000,
                "fte": 1,
            }
        ],
        "non_labor_lines": [],
    }


class CrValidatorTests(unittest.TestCase):
    def test_conforming_fixture_passes(self):
        result = self._run(fixture_workbook(), fixture_payload())

        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        self.assertEqual(json.loads(result.stdout)["status"], "pass")

    def test_price_labeled_total_equal_to_cost_subtotal_fails(self):
        workbook = fixture_workbook()
        workbook["IGCE Summary"]["B23"] = 1000000

        result = self._run(workbook, fixture_payload())

        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        failures = json.loads(result.stdout)["failures"]
        self.assertTrue(
            any("must equal estimated cost" in failure for failure in failures),
            failures,
        )

    def test_scenario_fee_base_drift_from_summary_fails(self):
        workbook = fixture_workbook()
        workbook["Scenario Analysis"]["D4"] = 1080000

        result = self._run(workbook, fixture_payload())

        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        failures = json.loads(result.stdout)["failures"]
        self.assertTrue(
            any("same fee-base rule" in failure for failure in failures),
            failures,
        )

    def test_multi_period_without_per_period_breakdown_fails(self):
        workbook = fixture_workbook()
        workbook["IGCE Summary"]["B19"] = 3
        payload = fixture_payload()
        payload["assumptions"]["periods"] = 3

        result = self._run(workbook, payload)

        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        failures = json.loads(result.stdout)["failures"]
        self.assertTrue(
            any("no per-period breakdown" in failure for failure in failures),
            failures,
        )

    def test_multi_period_with_per_period_breakdown_passes(self):
        workbook = fixture_workbook()
        summary = workbook["IGCE Summary"]
        summary["B19"] = 3
        summary["A25"] = "Base Year"
        summary["A26"] = "Option Year 1"
        summary["A27"] = "Option Year 2"
        payload = fixture_payload()
        payload["assumptions"]["periods"] = 3

        result = self._run(workbook, payload)

        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        self.assertEqual(json.loads(result.stdout)["status"], "pass")

    def _run(self, workbook: Workbook, payload: dict) -> subprocess.CompletedProcess[str]:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "fixture.xlsx"
            payload_path = root / "expected.json"
            workbook.save(workbook_path)
            payload_path.write_text(json.dumps(payload), encoding="utf-8")
            return subprocess.run(
                [
                    sys.executable,
                    str(VALIDATOR),
                    str(workbook_path),
                    "--expected",
                    str(payload_path),
                    "--engine",
                    "none",
                    "--json",
                ],
                capture_output=True,
                text=True,
                check=False,
            )


class TextClippingAuditTests(unittest.TestCase):
    """A text cell overflows only into an empty neighbour; anything else clips."""

    @classmethod
    def setUpClass(cls):
        cls.validator = load_clipping_validator("igce-builder-cr")

    def _sheet(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Scenario Analysis"
        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 12
        return workbook, sheet

    def test_long_label_beside_occupied_neighbour_fails(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["B1"] = "SOC 15-1252"

        failures = self.validator.text_clipping_audit(workbook)

        self.assertEqual(len(failures), 1, failures)
        self.assertIn("Scenario Analysis!A1 is clipped in print", failures[0])
        self.assertIn("Scenario Analysis: Senior Software Engineer", failures[0])
        self.assertIn("column A gives 12", failures[0])
        self.assertIn("B1 blocks the overflow", failures[0])
        self.assertIn("widen the column to at least 36", failures[0])

    def test_long_label_with_empty_neighbour_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_blank_string_neighbour_counts_as_empty(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["B1"] = "   "

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_merged_label_that_fits_the_span_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["D1"] = "SOC 15-1252"
        sheet.column_dimensions["A"].width = 44
        sheet.merge_cells("A1:C1")

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_wrapped_cell_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["A1"].alignment = Alignment(wrap_text=True)
        sheet["B1"] = "SOC 15-1252"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_label_that_fits_its_column_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Senior SWE"
        sheet["B1"] = "SOC 15-1252"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_borderline_label_within_tolerance_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "abcdefghijklm"
        sheet["B1"] = "occupied"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_right_aligned_label_overflows_left_not_right(self):
        workbook, sheet = self._sheet()
        sheet["B1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["B1"].alignment = Alignment(horizontal="right")
        sheet["C1"] = "occupied"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

        sheet["A1"] = "occupied"
        failures = self.validator.text_clipping_audit(workbook)
        self.assertEqual(len(failures), 1, failures)
        self.assertIn("A1 blocks the overflow", failures[0])

    def test_formula_and_numeric_cells_are_not_audited(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "=CONCATENATE(B1,C1,\"a very long generated label indeed\")"
        sheet["B1"] = 1234567890123456789
        sheet["C1"] = "occupied"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_bold_text_is_measured_wider(self):
        workbook, sheet = self._sheet()
        sheet.column_dimensions["A"].width = 20
        sheet["A1"] = "Period Totals (Current)"
        sheet["B1"] = "occupied"
        plain = self.validator.estimated_text_width(sheet["A1"].value, sheet["A1"].font)
        bold = self.validator.estimated_text_width(sheet["A1"].value, Font(bold=True))

        self.assertGreater(bold, plain)
        self.assertAlmostEqual(bold / plain, self.validator.CLIPPING_BOLD_FACTOR, places=6)

    def test_structural_audit_reports_clipping(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["B1"] = "SOC 15-1252"

        failures = self.validator.structural_audit(workbook, {"required_sheets": []})

        self.assertTrue(
            any("Scenario Analysis!A1 is clipped in print" in failure for failure in failures),
            failures,
        )

    def test_conforming_fixture_passes_every_remedy(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Scenario Analysis"
        sheet.column_dimensions["A"].width = 46
        sheet.column_dimensions["B"].width = 16
        sheet.column_dimensions["C"].width = 16
        # Merged block title spanning the block.
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet.merge_cells("A1:C1")
        # Label that fits its own column beside an occupied neighbour.
        sheet["A2"] = "Aged Annual Wage"
        sheet["B2"] = 183633.94
        # Wrapped narrative beside an occupied neighbour.
        sheet["A3"] = "Directional only: the CALC+ pool holds 28 records, below the guidance threshold."
        sheet["A3"].alignment = Alignment(wrap_text=True)
        sheet.row_dimensions[3].height = 60
        sheet["B3"] = "CALC+"
        # Long label on a row whose neighbours are empty.
        sheet["A4"] = "Government share plus performer contribution less project cost"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])


if __name__ == "__main__":
    unittest.main()
