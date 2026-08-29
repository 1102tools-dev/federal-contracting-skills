from __future__ import annotations

import importlib.util
import json
import subprocess
import sys
import tempfile
import unittest
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


ROOT = Path(__file__).resolve().parents[1]
VALIDATOR = ROOT / "skills" / "igce-builder-lh-tm" / "scripts" / "validate_workbook.py"


def load_clipping_validator(skill: str):
    directory = ROOT / "skills" / skill / "scripts"
    spec = importlib.util.spec_from_file_location(
        f"{skill.replace('-', '_')}_clipping_tests",
        directory / "validate_workbook.py",
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.path.insert(0, str(directory))
    try:
        spec.loader.exec_module(module)
    finally:
        sys.path.pop(0)
    return module


def autosize_for_print(workbook: Workbook, skill: str) -> Workbook:
    """Set every column width from the longest label actually written to it.

    This is the technique workbook-specification.md now requires of the
    generators: text cannot overflow into an occupied neighbour, so the column
    must be wide enough for its own widest label.
    """
    module = load_clipping_validator(skill)
    for sheet in workbook.worksheets:
        needed: dict[str, float] = {}
        for row in sheet.iter_rows():
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value.strip() or value.startswith("="):
                    continue
                width = module.estimated_text_width(value, cell.font)
                letter = cell.column_letter
                needed[letter] = max(needed.get(letter, 0.0), width)
        for letter, width in needed.items():
            current = sheet.column_dimensions[letter].width or 0.0
            sheet.column_dimensions[letter].width = max(current, min(90.0, width + 2.0))
    return workbook


def fixture_workbook(handling_value=0) -> Workbook:
    workbook = Workbook()
    workbook.active.title = "IGCE Summary"
    for name in (
        "Scenario Analysis",
        "Rate Validation",
        "Travel Detail",
        "Materials Detail",
        "Methodology",
        "Raw Data",
    ):
        workbook.create_sheet(name)

    summary = workbook["IGCE Summary"]
    summary["B2"] = 1.8
    summary["B3"] = 2.0
    summary["B4"] = 2.2
    summary["B5"] = 0.025
    summary["B8"] = "2026-01"
    summary["B9"] = "2026-01"
    summary["B10"] = (
        "=MAX(0,(VALUE(LEFT(B9,4))-VALUE(LEFT(B8,4)))*12+"
        "VALUE(MID(B9,6,2))-VALUE(MID(B8,6,2)))"
    )
    summary["B11"] = "=(1+B5)^(B10/12)"
    summary["B11"].number_format = "0.0000"

    scenario = workbook["Scenario Analysis"]
    scenario["A1"] = "Scenario Analysis: Test Analyst"
    scenario["B2"] = 100000
    scenario["B3"] = "='IGCE Summary'!$B$11"
    scenario["B3"].number_format = "0.0000"
    scenario["B4"] = "=B2*B3"
    scenario["B5"] = "=B4/2080"
    scenario["B7"] = "='IGCE Summary'!$B$2"
    scenario["B8"] = "=B5*B7"
    scenario["B9"] = "='IGCE Summary'!$B$3"
    scenario["B10"] = "=B5*B9"
    scenario["B11"] = "='IGCE Summary'!$B$4"
    scenario["B12"] = "=B5*B11"

    materials = workbook["Materials Detail"]
    materials["A1"] = "Item"
    materials["G1"] = "Material Handling Indirect"
    materials["A2"] = "Cloud hosting"
    materials["G2"] = handling_value
    return autosize_for_print(workbook, "igce-builder-lh-tm")


def multi_period_workbook(handling_value=0) -> Workbook:
    workbook = fixture_workbook(handling_value)
    summary = workbook["IGCE Summary"]
    summary["A5"] = "Escalation Rate"
    summary["A14"] = "Total Periods (Base plus Options)"
    summary["B14"] = 3
    summary["A20"] = "Base Year total"
    summary["A21"] = "Option Year 1 total"
    summary["A22"] = "Option Year 2 total"
    return autosize_for_print(workbook, "igce-builder-lh-tm")


def scenario_period_workbook(
    mid_values=(2500000.0, 2562500.0, 2626562.5), zeroed=False
) -> Workbook:
    """Multi-period fixture with a Scenario Analysis period-totals table.

    The IGCE Summary per-period totals are always the nonzero mid_values; the
    scenario table carries the same values unless zeroed is set, which models a
    stale-formula workbook whose cached scenario results collapsed to zero.
    """
    workbook = multi_period_workbook()
    scenario_values = (0.0, 0.0, 0.0) if zeroed else mid_values

    summary = workbook["IGCE Summary"]
    summary["B20"] = mid_values[0]
    summary["B21"] = mid_values[1]
    summary["B22"] = mid_values[2]
    summary["A23"] = "TOTAL ALL PERIODS"
    summary["B23"] = sum(mid_values)

    register = workbook["Raw Data"]
    register["A1"] = "Per-period labor totals"
    for offset, amount in enumerate((*mid_values, sum(mid_values))):
        register.cell(row=2 + offset, column=1, value=amount)

    scenario = workbook["Scenario Analysis"]
    scenario["A20"] = "PERIOD TOTALS FROM THE BURDENED RATES ABOVE"
    scenario["A21"] = "Period"
    scenario["B21"] = "Labor (Mid)"
    scenario["A22"] = "Base Year"
    scenario["B22"] = scenario_values[0]
    scenario["A23"] = "Option Year 1"
    scenario["B23"] = scenario_values[1]
    scenario["A24"] = "Option Year 2"
    scenario["B24"] = scenario_values[2]
    scenario["A25"] = "TOTAL LABOR, ALL PERIODS"
    scenario["B25"] = sum(scenario_values)
    return autosize_for_print(workbook, "igce-builder-lh-tm")


def multi_period_payload() -> dict:
    payload = fixture_payload()
    payload["assumptions"]["periods"] = 3
    return payload


def fixture_payload() -> dict:
    return {
        "assumptions": {
            "contract_type": "T&M",
            "burden_low": 1.8,
            "burden_mid": 2.0,
            "burden_high": 2.2,
            "aging_factor": 1.0,
            "productive_hours": 1880,
        },
        "labor_lines": [
            {
                "name": "Test Analyst",
                "annual_wage": 100000,
                "fte": 1,
            }
        ],
        "non_labor_lines": [],
    }


class LhTmMaterialHandlingValidatorTests(unittest.TestCase):
    def test_zero_material_handling_passes_integrated_validator(self):
        result = self._run(fixture_workbook(0), fixture_payload())

        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        self.assertEqual(json.loads(result.stdout)["status"], "pass")

    def test_undisclosed_percentage_formula_fails_integrated_validator(self):
        result = self._run(fixture_workbook("=D2*E2*0.07"), fixture_payload())

        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        payload = json.loads(result.stdout)
        self.assertEqual(payload["status"], "fail")
        self.assertIn(
            "Materials Detail!G2 contains undisclosed material handling '=D2*E2*0.07'",
            payload["failures"][0],
        )

    def test_disclosed_formula_and_basis_pass_integrated_validator(self):
        payload = fixture_payload()
        payload["material_handling_assertions"] = [
            {
                "cell": "'Materials Detail'!G2",
                "equals": "=D2*E2*0.07",
                "basis": "User-supplied accounting practice dated 2026-08-23",
            }
        ]
        result = self._run(fixture_workbook("=D2*E2*0.07"), payload)

        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        self.assertEqual(json.loads(result.stdout)["status"], "pass")

    def test_multi_period_fixture_passes(self):
        result = self._run(multi_period_workbook(), multi_period_payload())

        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        self.assertEqual(json.loads(result.stdout)["status"], "pass")

    def test_multi_period_without_per_period_totals_fails(self):
        workbook = multi_period_workbook()
        summary = workbook["IGCE Summary"]
        for coordinate in ("A20", "A21", "A22"):
            summary[coordinate] = None

        result = self._run(workbook, multi_period_payload())

        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        failures = json.loads(result.stdout)["failures"]
        self.assertTrue(
            any("no per-period totals" in failure for failure in failures),
            failures,
        )

    def test_dead_escalation_input_fails(self):
        workbook = multi_period_workbook()
        summary = workbook["IGCE Summary"]
        summary["A5"] = None
        summary["A6"] = "Escalation Rate"
        summary["C6"] = 0.03

        result = self._run(workbook, multi_period_payload())

        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        failures = json.loads(result.stdout)["failures"]
        self.assertTrue(
            any("referenced by zero formulas" in failure for failure in failures),
            failures,
        )

    def test_undisclosed_summary_money_constant_fails(self):
        workbook = multi_period_workbook()
        summary = workbook["IGCE Summary"]
        summary["A25"] = "Other direct costs"
        summary["B25"] = 75000

        result = self._run(workbook, multi_period_payload())

        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        failures = json.loads(result.stdout)["failures"]
        self.assertTrue(
            any("Raw Data refresh register" in failure for failure in failures),
            failures,
        )

    def test_registered_summary_money_constant_passes(self):
        workbook = multi_period_workbook()
        summary = workbook["IGCE Summary"]
        summary["A25"] = "Other direct costs"
        summary["B25"] = 75000
        register = workbook["Raw Data"]
        register["A1"] = "ODCs"
        register["B1"] = "$75,000 vendor quote"

        result = self._run(workbook, multi_period_payload())

        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        self.assertEqual(json.loads(result.stdout)["status"], "pass")

    def test_scenario_period_totals_tied_to_summary_pass(self):
        result = self._run(scenario_period_workbook(), multi_period_payload())

        self.assertEqual(result.returncode, 0, result.stdout + result.stderr)
        self.assertEqual(json.loads(result.stdout)["status"], "pass")

    def test_zeroed_scenario_period_totals_fail(self):
        result = self._run(
            scenario_period_workbook(zeroed=True), multi_period_payload()
        )

        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        failures = json.loads(result.stdout)["failures"]
        self.assertTrue(
            any(
                "are all zero while IGCE Summary per-period totals are nonzero"
                in failure
                for failure in failures
            ),
            failures,
        )

    def test_scenario_period_total_off_by_more_than_one_dollar_fails(self):
        workbook = scenario_period_workbook()
        workbook["Scenario Analysis"]["B22"] = 2500005.0

        result = self._run(workbook, multi_period_payload())

        self.assertEqual(result.returncode, 1, result.stdout + result.stderr)
        failures = json.loads(result.stdout)["failures"]
        self.assertTrue(
            any(
                "no IGCE Summary amount for that period is within $1" in failure
                for failure in failures
            ),
            failures,
        )

    def _run(self, workbook: Workbook, payload: dict) -> subprocess.CompletedProcess[str]:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            workbook_path = root / "fixture.xlsx"
            payload_path = root / "expected.json"
            workbook.save(workbook_path)
            payload_path.write_text(json.dumps(payload), encoding="utf-8")
            return subprocess.run(
                [
                    sys.executable,
                    str(VALIDATOR),
                    str(workbook_path),
                    "--expected",
                    str(payload_path),
                    "--engine",
                    "none",
                    "--json",
                ],
                capture_output=True,
                text=True,
                check=False,
            )


class TextClippingAuditTests(unittest.TestCase):
    """A text cell overflows only into an empty neighbour; anything else clips."""

    @classmethod
    def setUpClass(cls):
        cls.validator = load_clipping_validator("igce-builder-lh-tm")

    def _sheet(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Scenario Analysis"
        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 12
        return workbook, sheet

    def test_long_label_beside_occupied_neighbour_fails(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["B1"] = "SOC 15-1252"

        failures = self.validator.text_clipping_audit(workbook)

        self.assertEqual(len(failures), 1, failures)
        self.assertIn("Scenario Analysis!A1 is clipped in print", failures[0])
        self.assertIn("Scenario Analysis: Senior Software Engineer", failures[0])
        self.assertIn("column A gives 12", failures[0])
        self.assertIn("B1 blocks the overflow", failures[0])
        self.assertIn("widen the column to at least 36", failures[0])

    def test_long_label_with_empty_neighbour_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_blank_string_neighbour_counts_as_empty(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["B1"] = "   "

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_merged_label_that_fits_the_span_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["D1"] = "SOC 15-1252"
        sheet.column_dimensions["A"].width = 44
        sheet.merge_cells("A1:C1")

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_wrapped_cell_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["A1"].alignment = Alignment(wrap_text=True)
        sheet["B1"] = "SOC 15-1252"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_label_that_fits_its_column_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Senior SWE"
        sheet["B1"] = "SOC 15-1252"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_borderline_label_within_tolerance_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "abcdefghijklm"
        sheet["B1"] = "occupied"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_right_aligned_label_overflows_left_not_right(self):
        workbook, sheet = self._sheet()
        sheet["B1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["B1"].alignment = Alignment(horizontal="right")
        sheet["C1"] = "occupied"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

        sheet["A1"] = "occupied"
        failures = self.validator.text_clipping_audit(workbook)
        self.assertEqual(len(failures), 1, failures)
        self.assertIn("A1 blocks the overflow", failures[0])

    def test_formula_and_numeric_cells_are_not_audited(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "=CONCATENATE(B1,C1,\"a very long generated label indeed\")"
        sheet["B1"] = 1234567890123456789
        sheet["C1"] = "occupied"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_bold_text_is_measured_wider(self):
        workbook, sheet = self._sheet()
        sheet.column_dimensions["A"].width = 20
        sheet["A1"] = "Period Totals (Current)"
        sheet["B1"] = "occupied"
        plain = self.validator.estimated_text_width(sheet["A1"].value, sheet["A1"].font)
        bold = self.validator.estimated_text_width(sheet["A1"].value, Font(bold=True))

        self.assertGreater(bold, plain)
        self.assertAlmostEqual(bold / plain, self.validator.CLIPPING_BOLD_FACTOR, places=6)

    def test_structural_audit_reports_clipping(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["B1"] = "SOC 15-1252"

        failures = self.validator.structural_audit(workbook, {"required_sheets": []})

        self.assertTrue(
            any("Scenario Analysis!A1 is clipped in print" in failure for failure in failures),
            failures,
        )

    def test_conforming_fixture_passes_every_remedy(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Scenario Analysis"
        sheet.column_dimensions["A"].width = 46
        sheet.column_dimensions["B"].width = 16
        sheet.column_dimensions["C"].width = 16
        # Merged block title spanning the block.
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet.merge_cells("A1:C1")
        # Label that fits its own column beside an occupied neighbour.
        sheet["A2"] = "Aged Annual Wage"
        sheet["B2"] = 183633.94
        # Wrapped narrative beside an occupied neighbour.
        sheet["A3"] = "Directional only: the CALC+ pool holds 28 records, below the guidance threshold."
        sheet["A3"].alignment = Alignment(wrap_text=True)
        sheet.row_dimensions[3].height = 60
        sheet["B3"] = "CALC+"
        # Long label on a row whose neighbours are empty.
        sheet["A4"] = "Government share plus performer contribution less project cost"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])


if __name__ == "__main__":
    unittest.main()
