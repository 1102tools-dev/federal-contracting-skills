from __future__ import annotations

import importlib.util
import sys
import tempfile
import unittest
from pathlib import Path
from unittest import mock

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font


ROOT = Path(__file__).resolve().parents[1]
SCRIPT_DIR = ROOT / "skills" / "ot-cost-analysis" / "scripts"


def load_validator():
    spec = importlib.util.spec_from_file_location(
        "ot_cost_validator_tests",
        SCRIPT_DIR / "validate_workbook.py",
    )
    module = importlib.util.module_from_spec(spec)
    assert spec.loader is not None
    sys.path.insert(0, str(SCRIPT_DIR))
    try:
        spec.loader.exec_module(module)
    finally:
        sys.path.pop(0)
    return module


class OtCachedErrorAuditTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.validator = load_validator()

    def test_cached_error_audit_reports_every_sheet_and_cell(self):
        workbook = Workbook()
        workbook.active.title = "Summary"
        workbook["Summary"]["Z99"] = "#VALUE!"
        detail = workbook.create_sheet("Detail")
        detail["B4"] = "#DIV/0!"

        self.assertEqual(
            self.validator.cached_error_audit(workbook),
            [
                "Summary!Z99 has cached error #VALUE!",
                "Detail!B4 has cached error #DIV/0!",
            ],
        )

    def test_integrated_run_rejects_unmapped_cached_error(self):
        result = self._run_with_calculated_value("#VALUE!")

        self.assertEqual(result["status"], "fail")
        self.assertIn("Calculated!Z99 has cached error #VALUE!", result["failures"])

    def test_integrated_run_accepts_clean_recalculated_workbook(self):
        result = self._run_with_calculated_value(42.0)

        self.assertEqual(result["status"], "pass", result["failures"])

    def _run_with_calculated_value(self, value):
        formula_workbook = Workbook()
        formula_workbook.active["A1"] = "=1+1"
        calculated_workbook = Workbook()
        calculated_workbook.active.title = "Calculated"
        calculated_workbook["Calculated"]["Z99"] = value

        with tempfile.TemporaryDirectory() as directory:
            workbook_path = Path(directory) / "fixture.xlsx"
            expected_path = Path(directory) / "expected.json"
            formula_workbook.save(workbook_path)
            expected_path.write_text("{}", encoding="utf-8")
            recalculated_temp = tempfile.TemporaryDirectory()
            self.addCleanup(recalculated_temp.cleanup)

            with (
                mock.patch.object(self.validator, "load_payload", return_value={}),
                mock.patch.object(self.validator, "calculate", return_value={}),
                mock.patch.object(self.validator, "structural_audit", return_value=[]),
                mock.patch.object(self.validator, "find_soffice", return_value=Path("/fake/soffice")),
                mock.patch.object(
                    self.validator,
                    "recalculate",
                    return_value=(recalculated_temp, Path(directory) / "recalculated.xlsx"),
                ),
                mock.patch.object(
                    self.validator,
                    "load_workbook",
                    side_effect=[formula_workbook, calculated_workbook],
                ),
                mock.patch.object(self.validator, "compare", return_value=[]),
            ):
                return self.validator.run(workbook_path, expected_path, "auto", 0.01)


from openpyxl.styles import Alignment
from openpyxl.worksheet.properties import PageSetupProperties


def _workbook_with_detail():
    workbook = Workbook()
    workbook.active.title = "OT Cost Summary"
    detail = workbook.create_sheet("Milestone Detail")
    benchmarks = workbook.create_sheet("Labor Benchmarking")
    benchmarks["A3"] = "Systems engineer"
    benchmarks["A4"] = "Cybersecurity specialist"
    return workbook, detail, benchmarks


class OtLaborBenchmarkAuditTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.validator = load_validator()

    def test_silent_rate_reuse_without_proxy_fails(self):
        workbook, detail, _ = _workbook_with_detail()
        detail["A5"] = "Cybersecurity specialist"
        detail["C5"] = "='Labor Benchmarking'!H3"
        detail["D5"] = "Prior bounded source benchmark"

        failures = self.validator.labor_benchmark_audit(workbook)

        self.assertEqual(len(failures), 1)
        self.assertIn("Cybersecurity specialist", failures[0])
        self.assertIn("Systems engineer", failures[0])

    def test_matching_benchmark_row_passes(self):
        workbook, detail, _ = _workbook_with_detail()
        detail["A5"] = "Systems engineer"
        detail["C5"] = "='Labor Benchmarking'!H3"
        detail["D5"] = "Prior bounded source benchmark"

        self.assertEqual(self.validator.labor_benchmark_audit(workbook), [])

    def test_named_proxy_basis_passes(self):
        workbook, detail, _ = _workbook_with_detail()
        detail["A5"] = "Test and evaluation lead"
        detail["C5"] = "='Labor Benchmarking'!H3"
        detail["D5"] = "Systems engineer benchmark used as proxy; no separate SOC pool"

        self.assertEqual(self.validator.labor_benchmark_audit(workbook), [])


class OtHoursReconciliationAuditTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.validator = load_validator()

    def _detail_with_repeated_hours(self):
        workbook, detail, _ = _workbook_with_detail()
        for row in (5, 16):
            detail.cell(row=row, column=1, value="Systems engineer")
            detail.cell(row=row, column=2, value=900)
            detail.cell(row=row, column=3, value="='Labor Benchmarking'!H3")
        return workbook, detail

    def test_identical_hours_without_note_fails(self):
        workbook, _ = self._detail_with_repeated_hours()

        failures = self.validator.hours_reconciliation_audit(workbook)

        self.assertEqual(len(failures), 1)
        self.assertIn("identical hours (900)", failures[0])
        self.assertIn("systems engineer", failures[0])

    def test_hours_basis_note_passes(self):
        workbook, detail = self._detail_with_repeated_hours()
        detail["D5"] = "Hours basis: 1.5 FTE x 6 weeks"

        self.assertEqual(self.validator.hours_reconciliation_audit(workbook), [])

    def test_varying_hours_pass(self):
        workbook, detail = self._detail_with_repeated_hours()
        detail["B16"] = 1200

        self.assertEqual(self.validator.hours_reconciliation_audit(workbook), [])


class OtNarrativeFormatAuditTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.validator = load_validator()

    def _summary_with_description(self, *, width, wrap):
        workbook = Workbook()
        summary = workbook.active
        summary.title = "OT Cost Summary"
        summary["B18"] = "Description"
        summary["B19"] = "Edge sensing prototype and cyber readiness"
        summary.column_dimensions["B"].width = width
        if wrap:
            summary["B19"].alignment = Alignment(wrap_text=True)
        return workbook

    def test_missing_wrap_fails(self):
        workbook = self._summary_with_description(width=32, wrap=False)

        failures = self.validator.narrative_format_audit(workbook)

        self.assertEqual(len(failures), 1)
        self.assertIn("does not have wrap text enabled", failures[0])

    def test_narrow_column_fails(self):
        workbook = self._summary_with_description(width=12, wrap=True)

        failures = self.validator.narrative_format_audit(workbook)

        self.assertEqual(len(failures), 1)
        self.assertIn("below the 28 floor", failures[0])

    def test_wrapped_wide_column_passes(self):
        workbook = self._summary_with_description(width=32, wrap=True)

        self.assertEqual(self.validator.narrative_format_audit(workbook), [])


class OtPrintSetupAuditTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.validator = load_validator()

    def _canonical_workbook(self, *, print_ready):
        workbook = Workbook()
        workbook.active.title = "OT Cost Summary"
        for sheet_name in self.validator.REQUIRED_SHEETS[1:]:
            workbook.create_sheet(sheet_name)
        for sheet_name in self.validator.REQUIRED_SHEETS:
            sheet = workbook[sheet_name]
            sheet["A1"] = f"{sheet_name} title"
            sheet["F4"] = "content"
            if print_ready:
                sheet.print_area = "A1:F10"
                sheet.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
                sheet.page_setup.orientation = "landscape"
        return workbook

    def test_missing_print_area_and_fit_to_page_fails(self):
        workbook = self._canonical_workbook(print_ready=False)

        failures = self.validator.print_setup_audit(workbook)

        self.assertIn("OT Cost Summary has no print area set over the populated range", failures)
        self.assertIn("OT Cost Summary does not enable fitToPage scaling", failures)
        self.assertEqual(len(failures), 2 * len(self.validator.REQUIRED_SHEETS))

    def test_wide_sheet_in_portrait_fails(self):
        workbook = self._canonical_workbook(print_ready=True)
        detail = workbook["Milestone Detail"]
        detail.page_setup.orientation = "portrait"
        detail.cell(row=4, column=11, value="Rate")

        failures = self.validator.print_setup_audit(workbook)

        self.assertEqual(len(failures), 1)
        self.assertIn("Milestone Detail uses 11 columns in portrait orientation", failures[0])
        self.assertIn("must print landscape", failures[0])

    def test_conforming_print_setup_passes(self):
        workbook = self._canonical_workbook(print_ready=True)

        self.assertEqual(self.validator.print_setup_audit(workbook), [])

    def test_narrow_sheet_in_portrait_passes(self):
        workbook = self._canonical_workbook(print_ready=True)
        workbook["Methodology"].page_setup.orientation = "portrait"

        self.assertEqual(self.validator.print_setup_audit(workbook), [])


class OtRecostAuditTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.validator = load_validator()

    def _recost_workbook(self, *, decomposed):
        workbook = Workbook()
        summary = workbook.active
        summary.title = "OT Cost Summary"
        summary["A1"] = "Recosting Decision Book"
        detail = workbook.create_sheet("Milestone Detail")
        headers = ["Change", "Milestone", "Cost element", "Basis", "Delta"]
        if decomposed:
            headers += ["Hours", "Rate"]
        for column, header in enumerate(headers, start=1):
            detail.cell(row=4, column=column, value=header)
        detail["A5"] = "C-01"
        detail["B5"] = "M2"
        detail["C5"] = "Cyber/readiness labor"
        detail["D5"] = "Additional integration and scans"
        if decomposed:
            detail["F5"] = 320
            detail["G5"] = 200.0
            detail["E5"] = "=F5*G5"
        else:
            detail["E5"] = 64000
        benchmarks = workbook.create_sheet("Labor Benchmarking")
        benchmarks["A4"] = "Role"
        benchmarks["A5"] = "Cyber/readiness labor"
        return workbook, detail, benchmarks

    def test_is_recost_workbook_detection(self):
        recost, _, _ = self._recost_workbook(decomposed=True)
        plain = Workbook()
        plain.active["A1"] = "Independent Prototype Cost Model"

        self.assertTrue(self.validator.is_recost_workbook(recost))
        self.assertFalse(self.validator.is_recost_workbook(plain))

    def test_orphan_benchmark_role_fails(self):
        workbook, _, benchmarks = self._recost_workbook(decomposed=True)
        benchmarks["A6"] = "UX/Accessibility Specialist"

        failures = self.validator.recost_audit(workbook, {})

        self.assertEqual(len(failures), 1)
        self.assertIn("UX/Accessibility Specialist", failures[0])
        self.assertIn("appears in no Milestone Detail row", failures[0])

    def test_lump_sum_labor_delta_fails(self):
        workbook, _, _ = self._recost_workbook(decomposed=False)

        failures = self.validator.recost_audit(workbook, {})

        self.assertEqual(len(failures), 1)
        self.assertIn("lump sum", failures[0])
        self.assertIn("hours x rate", failures[0])

    def test_missing_register_element_fails(self):
        workbook, _, _ = self._recost_workbook(decomposed=True)
        payload = {"recost_register_elements": ["labor", "travel"]}

        failures = self.validator.recost_audit(workbook, payload)

        self.assertEqual(len(failures), 1)
        self.assertIn("'travel'", failures[0])
        self.assertIn("$0 delta", failures[0])

    def test_conforming_recost_passes(self):
        workbook, detail, _ = self._recost_workbook(decomposed=True)
        detail["A6"] = "C-02"
        detail["B6"] = "M2"
        detail["C6"] = "Travel"
        detail["D6"] = "Register directed repricing; no trip changes, so zero delta"
        detail["E6"] = 0
        payload = {"recost_register_elements": ["labor", "travel"]}

        self.assertEqual(self.validator.recost_audit(workbook, payload), [])

    def test_invalid_register_elements_payload_raises(self):
        workbook, _, _ = self._recost_workbook(decomposed=True)

        with self.assertRaises(self.validator.InputError):
            self.validator.recost_audit(workbook, {"recost_register_elements": "travel"})


class OtHoursDerivationAuditTests(unittest.TestCase):
    @classmethod
    def setUpClass(cls):
        cls.validator = load_validator()

    def _hours_workbook(
        self,
        *,
        hours="derived",
        drivers=True,
        reconciliation="formula",
        scenario=None,
    ):
        workbook = Workbook()
        workbook.active.title = "OT Cost Summary"
        detail = workbook.create_sheet("Milestone Detail")
        detail["A2"] = "M1: baseline and test design"
        detail["A3"], detail["B3"] = "Duration (weeks)", 6
        headers = ["Labor Category", "FTE Loading", "Weeks", "Hours per FTE-week", "Hours"]
        if not drivers:
            headers = ["Labor Category", "Hours"]
        for column, header in enumerate(headers, start=1):
            detail.cell(row=4, column=column, value=header)
        hours_column = len(headers)
        for offset, (name, fte) in enumerate((("Systems engineer", 1.0), ("Software developer", 0.5))):
            row = 5 + offset
            detail.cell(row=row, column=1, value=name)
            if drivers:
                detail.cell(row=row, column=2, value=fte)
                detail.cell(row=row, column=3, value=6)
                detail.cell(row=row, column=4, value=36.1538)
            if hours == "derived":
                value = f"=B{row}*C{row}*D{row}"
            elif hours == "unlinked":
                value = f"={fte}*6*36.1538"
            else:
                value = round(fte * 6 * 36.1538, 4)
            detail.cell(row=row, column=hours_column, value=value)
        detail.cell(row=7, column=1, value="Labor subtotal")
        detail.cell(row=7, column=hours_column, value="=SUM(E5:E6)")
        if reconciliation == "formula":
            detail.cell(row=8, column=1, value="Hours reconcile to duration and staffing")
            detail.cell(row=8, column=2, value='=IF(ROUND(E7-SUM(B5:B6)*B3*D5,4)=0,"OK","MISMATCH")')
        elif reconciliation == "prose":
            detail.cell(row=8, column=1, value="Hours basis: 1.5 FTE x 6 weeks. Hours reconcile to M1.")
        if scenario is not None:
            analysis = workbook.create_sheet("Scenario Analysis")
            analysis["A4"], analysis["B4"] = "Labor category", "M1 hours"
            analysis["A5"] = "Systems engineer"
            analysis["B5"] = 216.9228 if scenario == "literal" else "='Milestone Detail'!E5"
        return workbook, detail

    def test_asserted_hours_constant_fails(self):
        workbook, _ = self._hours_workbook(hours="asserted")

        failures = self.validator.hours_derivation_audit(workbook)

        self.assertEqual(len(failures), 1)
        self.assertIn("Milestone Detail!E5 asserts hours for 'Systems engineer'", failures[0])
        self.assertIn("2 of 2 labor rows", failures[0])
        self.assertIn("must be a formula", failures[0])

    def test_missing_driver_columns_fail(self):
        workbook, _ = self._hours_workbook(hours="asserted", drivers=False)

        failures = self.validator.hours_derivation_audit(workbook)

        self.assertEqual(len(failures), 2)
        self.assertIn("exposes no FTE loading, weeks or duration, hours per FTE-week", failures[0])
        self.assertIn("asserts hours", failures[1])

    def test_hours_formula_ignoring_driver_cells_fails(self):
        workbook, _ = self._hours_workbook(hours="unlinked")

        failures = self.validator.hours_derivation_audit(workbook)

        self.assertEqual(len(failures), 6)
        self.assertIn("Milestone Detail!E5 hours formula does not reference its FTE loading", failures[0])
        self.assertIn("driver cell B5", failures[0])

    def test_non_numeric_driver_cell_fails(self):
        workbook, detail = self._hours_workbook()
        detail["B5"] = "1.0 FTE"

        failures = self.validator.hours_derivation_audit(workbook)

        self.assertEqual(len(failures), 1)
        self.assertIn("Milestone Detail!B5 FTE loading driver is not a numeric input cell", failures[0])

    def test_prose_reconciliation_fails(self):
        workbook, _ = self._hours_workbook(reconciliation="prose")

        failures = self.validator.hours_derivation_audit(workbook)

        self.assertEqual(len(failures), 1)
        self.assertIn("carry no formula reconciling derived hours", failures[0])
        self.assertIn("OK or MISMATCH state", failures[0])
        self.assertIn("'Hours basis:' note is prose", failures[0])

    def test_missing_reconciliation_fails(self):
        workbook, _ = self._hours_workbook(reconciliation=None)

        failures = self.validator.hours_derivation_audit(workbook)

        self.assertEqual(len(failures), 1)
        self.assertIn("carry no formula reconciling derived hours", failures[0])
        self.assertNotIn("Hours basis", failures[0])

    def test_scenario_sheet_restating_hours_as_literals_fails(self):
        workbook, _ = self._hours_workbook(scenario="literal")

        failures = self.validator.hours_derivation_audit(workbook)

        self.assertEqual(len(failures), 1)
        self.assertIn("Scenario Analysis!B5 restates milestone hours as the constant", failures[0])
        self.assertIn("must reference those cells by formula", failures[0])

    def test_conforming_derived_hours_pass(self):
        workbook, _ = self._hours_workbook(scenario="linked")

        self.assertEqual(self.validator.hours_derivation_audit(workbook), [])

    def test_workbook_without_labor_block_is_not_audited(self):
        workbook = Workbook()
        workbook.active.title = "OT Cost Summary"
        workbook.create_sheet("Milestone Detail")["A1"] = "Milestone Detail"

        self.assertEqual(self.validator.hours_derivation_audit(workbook), [])

    def test_structural_audit_wires_in_the_hours_derivation_gate(self):
        workbook, _ = self._hours_workbook(hours="asserted")

        failures = self.validator.structural_audit(workbook, {"required_sheets": []})

        self.assertTrue(
            any("asserts hours for 'Systems engineer'" in failure for failure in failures),
            failures,
        )


class TextClippingAuditTests(unittest.TestCase):
    """A text cell overflows only into an empty neighbour; anything else clips."""

    @classmethod
    def setUpClass(cls):
        cls.validator = load_validator()

    def _sheet(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Scenario Analysis"
        sheet.column_dimensions["A"].width = 12
        sheet.column_dimensions["B"].width = 12
        sheet.column_dimensions["C"].width = 12
        return workbook, sheet

    def test_long_label_beside_occupied_neighbour_fails(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["B1"] = "SOC 15-1252"

        failures = self.validator.text_clipping_audit(workbook)

        self.assertEqual(len(failures), 1, failures)
        self.assertIn("Scenario Analysis!A1 is clipped in print", failures[0])
        self.assertIn("Scenario Analysis: Senior Software Engineer", failures[0])
        self.assertIn("column A gives 12", failures[0])
        self.assertIn("B1 blocks the overflow", failures[0])
        self.assertIn("widen the column to at least 36", failures[0])

    def test_long_label_with_empty_neighbour_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_blank_string_neighbour_counts_as_empty(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["B1"] = "   "

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_merged_label_that_fits_the_span_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["D1"] = "SOC 15-1252"
        sheet.column_dimensions["A"].width = 44
        sheet.merge_cells("A1:C1")

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_wrapped_cell_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["A1"].alignment = Alignment(wrap_text=True)
        sheet["B1"] = "SOC 15-1252"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_label_that_fits_its_column_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Senior SWE"
        sheet["B1"] = "SOC 15-1252"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_borderline_label_within_tolerance_passes(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "abcdefghijklm"
        sheet["B1"] = "occupied"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_right_aligned_label_overflows_left_not_right(self):
        workbook, sheet = self._sheet()
        sheet["B1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["B1"].alignment = Alignment(horizontal="right")
        sheet["C1"] = "occupied"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

        sheet["A1"] = "occupied"
        failures = self.validator.text_clipping_audit(workbook)
        self.assertEqual(len(failures), 1, failures)
        self.assertIn("A1 blocks the overflow", failures[0])

    def test_formula_and_numeric_cells_are_not_audited(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "=CONCATENATE(B1,C1,\"a very long generated label indeed\")"
        sheet["B1"] = 1234567890123456789
        sheet["C1"] = "occupied"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])

    def test_bold_text_is_measured_wider(self):
        workbook, sheet = self._sheet()
        sheet.column_dimensions["A"].width = 20
        sheet["A1"] = "Period Totals (Current)"
        sheet["B1"] = "occupied"
        plain = self.validator.estimated_text_width(sheet["A1"].value, sheet["A1"].font)
        bold = self.validator.estimated_text_width(sheet["A1"].value, Font(bold=True))

        self.assertGreater(bold, plain)
        self.assertAlmostEqual(bold / plain, self.validator.CLIPPING_BOLD_FACTOR, places=6)

    def test_structural_audit_reports_clipping(self):
        workbook, sheet = self._sheet()
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet["B1"] = "SOC 15-1252"

        failures = self.validator.structural_audit(workbook, {"required_sheets": []})

        self.assertTrue(
            any("Scenario Analysis!A1 is clipped in print" in failure for failure in failures),
            failures,
        )

    def test_conforming_fixture_passes_every_remedy(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Scenario Analysis"
        sheet.column_dimensions["A"].width = 46
        sheet.column_dimensions["B"].width = 16
        sheet.column_dimensions["C"].width = 16
        # Merged block title spanning the block.
        sheet["A1"] = "Scenario Analysis: Senior Software Engineer"
        sheet.merge_cells("A1:C1")
        # Label that fits its own column beside an occupied neighbour.
        sheet["A2"] = "Aged Annual Wage"
        sheet["B2"] = 183633.94
        # Wrapped narrative beside an occupied neighbour.
        sheet["A3"] = "Directional only: the CALC+ pool holds 28 records, below the guidance threshold."
        sheet["A3"].alignment = Alignment(wrap_text=True)
        sheet.row_dimensions[3].height = 60
        sheet["B3"] = "CALC+"
        # Long label on a row whose neighbours are empty.
        sheet["A4"] = "Government share plus performer contribution less project cost"

        self.assertEqual(self.validator.text_clipping_audit(workbook), [])


if __name__ == "__main__":
    unittest.main()
